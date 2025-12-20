import asyncio
import atexit
import os
from datetime import datetime, timedelta, timezone
import re
import time
import requests
import aiohttp
import shlex
import urllib.parse
import sqlite3
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Optional, List, Dict

from aiogram import Bot, Dispatcher, F
from aiogram.exceptions import TelegramConflictError
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    BotCommand,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    BotCommandScopeChat,
    FSInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from dotenv import load_dotenv

import database as db
from utils.distance import haversine_m
from utils import notifications
from utils.reports import check_rules_violation, get_status_color, get_status_name
from utils.exports import generate_csv_report
from utils.geocoding import reverse_geocode, reverse_geocode_background
import csv

# ================== KONFİQURASİYA ==================

LOCK_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".bot_instance.lock")


def _is_pid_running(pid: int) -> bool:
    try:
        if pid <= 0:
            return False
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def acquire_single_instance_lock() -> bool:
    try:
        fd = os.open(LOCK_FILE_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        try:
            os.write(fd, str(os.getpid()).encode("utf-8"))
        finally:
            os.close(fd)
        return True
    except FileExistsError:
        try:
            with open(LOCK_FILE_PATH, "r", encoding="utf-8") as f:
                raw = (f.read() or "").strip()
            pid = int(raw) if raw.isdigit() else -1
            if pid != -1 and _is_pid_running(pid):
                return False
        except Exception:
            # If lock file is unreadable, treat it as stale and try to replace
            pass

        # Stale lock - remove and retry once
        try:
            os.remove(LOCK_FILE_PATH)
        except Exception:
            return False
        return acquire_single_instance_lock()


def release_single_instance_lock() -> None:
    try:
        os.remove(LOCK_FILE_PATH)
    except FileNotFoundError:
        pass

load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is not set in .env")

ADMIN_ID = int(os.getenv("ADMIN_ID", "0") or "0")

# Lokasiya gözləmə vaxtı (saniyə)
LOCATION_TIMEOUT = 120

# Telegram mesaj limitindən bir az aşağı saxlayırıq
TG_CHUNK_LIMIT = 3500

# Məkən koordinatları (lat, lon) və radius (metrlə)
WORKPLACE_LAT = float(os.getenv("WORKPLACE_LAT", "40.4093"))  # Baku koordinatları default
WORKPLACE_LON = float(os.getenv("WORKPLACE_LON", "49.8671"))
WORKPLACE_RADIUS_M = min(float(os.getenv("WORKPLACE_RADIUS_M", "300")), 300.0)

# Giriş-çıxış vaxt məhdudiyyətləri
CHECKIN_DEADLINE_HOUR = 11  # Giriş 11:00-a qədər
CHECKOUT_DEADLINE_HOUR = 19  # Çıxış 19:00-a qədər
MIN_WORK_DURATION_HOURS = 6  # Minimum iş müddəti (saat)

# Lokasiya dəqiqliyi (metrlə) - eyni yer sayılması üçün tolerance
LOCATION_TOLERANCE_M = 50  # 50 metr tolerance

try:
    BAKU_TZ = ZoneInfo("Asia/Baku")
except Exception:
    BAKU_TZ = timezone(timedelta(hours=4))


def now_baku() -> datetime:
    return datetime.now(BAKU_TZ)


def today_baku() -> str:
    return now_baku().date().isoformat()


def parse_dt_to_baku(value) -> datetime:
    if isinstance(value, datetime):
        dt = value
    else:
        dt = datetime.fromisoformat(str(value))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(BAKU_TZ)

# ================== GLOBAL OBYEKTLƏR ==================

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# In-memory pending action per user: (action, ts)
# action: "checkin" | "checkout", ts: unix timestamp
pending_action: dict[int, tuple[str, float]] = {}

PROFESSIONS: list[str] = [
    "Aşpaz",
    "Dərzi",
    "Qənnadı",
    "Qrafik dizayn",
    "Permanent",
    "Full stack",
    "Satıcı/kassir",
]


# ================== FSM STATES ==================

class Reg(StatesGroup):
    profession = State()
    code = State()
    name = State()
    fin = State()
    document_series_number = State()
    phone_number = State()


class AdminAddG(StatesGroup):
    profession = State()
    code = State()


class AdminReportByCode(StatesGroup):
    date = State()
    code = State()


class AdminExcelReport(StatesGroup):
    date = State()
    format_type = State()  # excel, csv, pdf


class AdminPeriodReport(StatesGroup):
    period_type = State()  # daily, weekly, monthly, code
    start_date = State()
    end_date = State()
    code = State()
    format_type = State()


class EditProfile(StatesGroup):
    field = State()
    new_value = State()

class AdminManageGroup(StatesGroup):
    action = State()  # create, add_code
    profession = State()
    code = State()
    date = State()


class AdminManageStudent(StatesGroup):
    action = State()  # list, delete, deactivate, activate
    code_or_id = State()
    field = State()
    new_value = State()
    confirm = State()


# ================== KÖMƏKÇİ FUNKSİYALAR ==================


def is_admin(user_id: int) -> bool:
    return ADMIN_ID != 0 and user_id == ADMIN_ID


def worker_keyboard() -> ReplyKeyboardMarkup:
    kb = [
        [KeyboardButton(text="🟢 Giriş"), KeyboardButton(text="🔴 Çıxış")],
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def admin_keyboard() -> ReplyKeyboardMarkup:
    kb = [
        [KeyboardButton(text="📊 Bu gün"), KeyboardButton(text="👥 İşçilər")],
        [KeyboardButton(text="➕ Bugünün kodu"), KeyboardButton(text="📜 Kodlar")],
        [KeyboardButton(text="🗒 Qeydiyyatlar"), KeyboardButton(text="📈 Kod üzrə hesabat")],
        [KeyboardButton(text="📥 Excel hesabat"), KeyboardButton(text="👨‍👩‍👧‍👦 Qruplar")],
        [KeyboardButton(text="🎓 Tələbələr"), KeyboardButton(text="🎛 Menyu")],
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def location_keyboard(prompt: str) -> ReplyKeyboardMarkup:
    kb = [[KeyboardButton(text=prompt, request_location=True)]]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True, one_time_keyboard=True)


def professions_keyboard() -> ReplyKeyboardMarkup:
    rows = []
    row: list[KeyboardButton] = []
    for i, p in enumerate(PROFESSIONS, start=1):
        row.append(KeyboardButton(text=f"{i}. {p}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([KeyboardButton(text="❌ Ləğv et")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def chunk_send(text: str):
    """Mesajı TG_CHUNK_LIMIT uzunluğunda parçalamaq üçün generator."""
    s = text
    while s:
        chunk = s[:TG_CHUNK_LIMIT]
        cut = chunk.rfind("\n")
        if cut == -1:
            cut = len(chunk)
        yield s[:cut]
        s = s[cut:].lstrip("\n")


def _match_profession(token: str) -> str | None:
    t = token.strip().lower()
    for p in PROFESSIONS:
        if p.lower() == t:
            return p
    return None


def _parse_date_or_today(s: str) -> str | None:
    """'YYYY-MM-DD' və ya 'bugun' tipini tarixə çevirir, yanlış olsa None qaytarır."""
    s = s.strip().lower()
    if s in ("bugun", "bu gun", "bu gün"):
        return today_baku()
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.date().isoformat()
    except ValueError:
        return None


async def schedule_checkout_reminder(telegram_id: int, checkin_time: datetime, user2_id: int) -> None:
    """Schedule a reminder to checkout after 8 hours from checkin time."""
    # Wait 8 hours (8 * 60 * 60 = 28800 seconds)
    wait_seconds = 8 * 60 * 60
    
    try:
        await asyncio.sleep(wait_seconds)
        
        # Check if user already checked out
        db.init_gps_tables()
        sess = db.get_open_session(user2_id)
        
        if sess:
            # User hasn't checked out yet, send reminder
            try:
                reminder_text = (
                    "⏰ Xatırlatma\n\n"
                    "8 saat keçib. Xahiş edirik çıxış edin.\n\n"
                    "🔴 Çıxış düyməsinə basın və lokasiyanızı göndərin."
                )
                await bot.send_message(telegram_id, reminder_text, reply_markup=worker_keyboard())
            except Exception as e:
                print(f"[schedule_checkout_reminder] Error sending reminder to {telegram_id}: {e}")
        # If session is closed, user already checked out, no need to remind
    except asyncio.CancelledError:
        # Task was cancelled, ignore
        pass
    except Exception as e:
        print(f"[schedule_checkout_reminder] Error: {e}")


def generate_daily_excel_report(date: str) -> str:
    """Generate Excel report for a specific date. Returns path to the Excel file."""
    # Get all users with attendance data for the date
    report_data = db.get_daily_report_for_excel(date)
    
    # Debug: print first few records to check data
    if report_data:
        print(f"[DEBUG] First record: {report_data[0]}")
        print(f"[DEBUG] giris_time type: {type(report_data[0].get('giris_time'))}, value: {report_data[0].get('giris_time')}")
        print(f"[DEBUG] cixis_time type: {type(report_data[0].get('cixis_time'))}, value: {report_data[0].get('cixis_time')}")
    
    # Group by code to show members
    by_code: dict[str, list[dict]] = {}
    for row in report_data:
        code = row.get('code') or '-'
        by_code.setdefault(code, []).append(row)
    
    # Calculate statistics
    total_workers = len(report_data)
    workers_with_giris = [r for r in report_data if r.get('giris_time')]
    workers_without_giris = [r for r in report_data if not r.get('giris_time')]
    workers_came = len(workers_with_giris)
    workers_not_came = len(workers_without_giris)
    
    # Qayda yoxlamaları statistikası
    ok_count = 0
    violation_count = 0
    inactive_count = 0
    
    for member in report_data:
        is_active = member.get('is_active', 1)
        if is_active == 0:
            inactive_count += 1
        else:
            giris_time = member.get('giris_time')
            cixis_time = member.get('cixis_time')
            start_lat = member.get('start_lat')
            start_lon = member.get('start_lon')
            end_lat = member.get('end_lat')
            end_lon = member.get('end_lon')
            
            status, _ = check_rules_violation(
                giris_time, cixis_time,
                float(start_lat) if start_lat is not None else None,
                float(start_lon) if start_lon is not None else None,
                float(end_lat) if end_lat is not None else None,
                float(end_lon) if end_lon is not None else None,
                is_active,
                CHECKIN_DEADLINE_HOUR, CHECKOUT_DEADLINE_HOUR, MIN_WORK_DURATION_HOURS,
                WORKPLACE_LAT, WORKPLACE_LON, WORKPLACE_RADIUS_M, LOCATION_TOLERANCE_M
            )
            if status == "ok":
                ok_count += 1
            elif status == "violation":
                violation_count += 1
    
    # Aktiv tələbə sayı
    active_count = db.get_active_students_count(date)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Hesabat {date}"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data style
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Format date for display (e.g., "15 Yanvar 2024")
    try:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        # Azerbaijani month names
        months_az = {
            1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
            5: "May", 6: "İyun", 7: "İyul", 8: "Avqust",
            9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
        }
        formatted_date = f"{date_obj.day} {months_az[date_obj.month]} {date_obj.year}"
    except:
        formatted_date = date
    
    # Write statistics section at the top
    stats_row = 1
    stats_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    stats_font = Font(bold=True, color="FFFFFF", size=14)
    
    ws.cell(row=stats_row, column=1, value="📊 STATİSTİKA").font = stats_font
    ws.cell(row=stats_row, column=1).fill = stats_fill
    ws.merge_cells(f'A{stats_row}:F{stats_row}')
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Aktiv tələbə sayı:").font = Font(bold=True, size=11)
    ws.cell(row=stats_row, column=2, value=active_count).font = Font(bold=True, size=11)
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Ümumi işçi sayı:").font = Font(bold=True, size=11)
    ws.cell(row=stats_row, column=2, value=total_workers).font = Font(bold=True, size=11)
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="İşə gələnlər:").font = Font(bold=True, size=11, color="006100")
    ws.cell(row=stats_row, column=2, value=workers_came).font = Font(bold=True, size=11, color="006100")
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="İşə gəlməyənlər:").font = Font(bold=True, size=11, color="C00000")
    ws.cell(row=stats_row, column=2, value=workers_not_came).font = Font(bold=True, size=11, color="C00000")
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Qaydalara uyğun:").font = Font(bold=True, size=11, color="006100")
    ws.cell(row=stats_row, column=2, value=ok_count).font = Font(bold=True, size=11, color="006100")
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Qayda pozuntusu:").font = Font(bold=True, size=11, color="C00000")
    ws.cell(row=stats_row, column=2, value=violation_count).font = Font(bold=True, size=11, color="C00000")
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Kursdan çıxarılan:").font = Font(bold=True, size=11, color="808080")
    ws.cell(row=stats_row, column=2, value=inactive_count).font = Font(bold=True, size=11, color="808080")
    
    # Write list of workers who didn't come
    if workers_without_giris:
        stats_row += 1
        ws.cell(row=stats_row, column=1, value="İşə gəlməyənlərin siyahısı:").font = Font(bold=True, size=11, color="C00000")
        stats_row += 1
        for worker in workers_without_giris:
            name = worker.get('name', '?')
            fin = worker.get('fin', '-')
            code = worker.get('code', '-')
            ws.cell(row=stats_row, column=1, value=f"• {name} (FIN: {fin}, Kod: {code})")
            stats_row += 1
    
    # Add empty row before main table
    stats_row += 1
    
    # Set column headers (genişləndirilmiş)
    headers = [
        "Tarix",
        "FIN Kodu",
        "Ad",
        "Soyad",
        "Vəsiqə Seriya",
        "Telefon",
        "Qrup Kodu",
        "Peşə",
        "Giriş Saatı",
        "Çıxış Saatı",
        "GPS Koordinatları",
        "Lokasiya",
        "Giriş Linki",
        "Çıxış Linki",
        "Status",
        "Qayda Pozuntuları",
        "Qrup Üzvləri"
    ]
    
    header_row = stats_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data (starting after header row)
    row_num = header_row + 1
    for code, members in sorted(by_code.items()):
        # Get all member names for this code
        member_names = [m.get('name', '?') for m in members]
        members_str = ", ".join(member_names) if member_names else "-"
        
        for member in members:
            name_full = member.get('name', '')
            # Split name into ad and soyad (assuming space-separated)
            name_parts = name_full.strip().split(maxsplit=1)
            ad = name_parts[0] if name_parts else name_full
            soyad = name_parts[1] if len(name_parts) > 1 else ""
            
            fin = member.get('fin', '')
            profession = member.get('profession', '-')
            seriya = member.get('seriya') or '-'
            phone_number = member.get('phone_number') or '-'
            is_active = member.get('is_active', 1)
            
            # Handle entry/exit times
            giris_time_raw = member.get('giris_time')
            cixis_time_raw = member.get('cixis_time')
            
            # Convert None to empty string for easier checking
            if giris_time_raw is None:
                giris_time_raw = ''
            else:
                giris_time_raw = str(giris_time_raw).strip()
            
            if cixis_time_raw is None:
                cixis_time_raw = ''
            else:
                cixis_time_raw = str(cixis_time_raw).strip()
            
            if not giris_time_raw:
                # No entry at all
                giris_time = "Giriş yoxdur"
                cixis_time = "Giriş yoxdur"
            elif not cixis_time_raw:
                # Has entry but no exit
                giris_time = giris_time_raw
                cixis_time = "Yoxdur"
            else:
                # Has both entry and exit
                giris_time = giris_time_raw
                cixis_time = cixis_time_raw
            
            # Handle address and maps link - prioritize GPS session coordinates
            # Get GPS session coordinates if available
            start_lat = member.get('start_lat')
            start_lon = member.get('start_lon')
            end_lat = member.get('end_lat')
            end_lon = member.get('end_lon')
            
            address = "-"
            start_link = "-"
            end_link = "-"
            gps_coords = "-"
            lat = None
            lon = None
            
            # Use GPS session coordinates if available (same as bugun report)
            if start_lat is not None and start_lon is not None:
                # Use start location (giris location)
                lat = float(start_lat)
                lon = float(start_lon)
                gps_coords = f"{lat}, {lon}"
                # Sync context: use coordinates (geocoding cache is async-only)
                address = gps_coords
                start_link = f"https://maps.google.com/?q={lat},{lon}"
            elif end_lat is not None and end_lon is not None:
                # Fallback to end location if start not available
                lat = float(end_lat)
                lon = float(end_lon)
                gps_coords = f"{lat}, {lon}"
                # Sync context: use coordinates (geocoding cache is async-only)
                address = gps_coords
                end_link = f"https://maps.google.com/?q={lat},{lon}"
            else:
                # Fallback to legacy attendance location fields
                giris_loc = member.get('giris_loc') or ''
                cixis_loc = member.get('cixis_loc') or ''
                
                if giris_loc:
                    address = str(giris_loc).strip()
                    # Generate Google Maps link from address
                    encoded_address = urllib.parse.quote(address)
                    start_link = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"
                elif cixis_loc:
                    address = str(cixis_loc).strip()
                    # Generate Google Maps link from address
                    encoded_address = urllib.parse.quote(address)
                    end_link = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"
            
            # Qayda yoxlaması
            status, violations = check_rules_violation(
                giris_time_raw if giris_time_raw else None,
                cixis_time_raw if cixis_time_raw else None,
                lat, lon,
                float(end_lat) if end_lat is not None else None,
                float(end_lon) if end_lon is not None else None,
                is_active,
                CHECKIN_DEADLINE_HOUR,
                CHECKOUT_DEADLINE_HOUR,
                MIN_WORK_DURATION_HOURS,
                WORKPLACE_LAT,
                WORKPLACE_LON,
                WORKPLACE_RADIUS_M,
                LOCATION_TOLERANCE_M
            )
            status_name = get_status_name(status)
            violations_str = "; ".join(violations) if violations else "-"
            
            # Rəng kodlaması
            status_color = get_status_color(status)
            row_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            
            # Write data with date in first column
            ws.cell(row=row_num, column=1, value=formatted_date).alignment = data_alignment
            ws.cell(row=row_num, column=1).fill = row_fill
            ws.cell(row=row_num, column=2, value=fin).alignment = data_alignment
            ws.cell(row=row_num, column=2).fill = row_fill
            ws.cell(row=row_num, column=3, value=ad).alignment = data_alignment
            ws.cell(row=row_num, column=3).fill = row_fill
            ws.cell(row=row_num, column=4, value=soyad).alignment = data_alignment
            ws.cell(row=row_num, column=4).fill = row_fill
            ws.cell(row=row_num, column=5, value=seriya).alignment = data_alignment
            ws.cell(row=row_num, column=5).fill = row_fill
            ws.cell(row=row_num, column=6, value=phone_number).alignment = data_alignment
            ws.cell(row=row_num, column=6).fill = row_fill
            ws.cell(row=row_num, column=7, value=code).alignment = data_alignment
            ws.cell(row=row_num, column=7).fill = row_fill
            ws.cell(row=row_num, column=8, value=profession).alignment = data_alignment
            ws.cell(row=row_num, column=8).fill = row_fill
            ws.cell(row=row_num, column=9, value=giris_time).alignment = data_alignment
            ws.cell(row=row_num, column=9).fill = row_fill
            ws.cell(row=row_num, column=10, value=cixis_time).alignment = data_alignment
            ws.cell(row=row_num, column=10).fill = row_fill
            ws.cell(row=row_num, column=11, value=gps_coords).alignment = data_alignment
            ws.cell(row=row_num, column=11).fill = row_fill
            ws.cell(row=row_num, column=12, value=address).alignment = data_alignment
            ws.cell(row=row_num, column=12).fill = row_fill
            
            # Add hyperlink for maps link if available
            if start_link != "-":
                cell = ws.cell(row=row_num, column=13, value="Giriş xəritə")
                cell.hyperlink = start_link
                cell.font = Font(color="0000FF", underline="single")
                cell.alignment = data_alignment
                cell.fill = row_fill
            else:
                ws.cell(row=row_num, column=13, value="-").alignment = data_alignment
                ws.cell(row=row_num, column=13).fill = row_fill

            if end_link != "-":
                cell2 = ws.cell(row=row_num, column=14, value="Çıxış xəritə")
                cell2.hyperlink = end_link
                cell2.font = Font(color="0000FF", underline="single")
                cell2.alignment = data_alignment
                cell2.fill = row_fill
            else:
                ws.cell(row=row_num, column=14, value="-").alignment = data_alignment
                ws.cell(row=row_num, column=14).fill = row_fill
            
            ws.cell(row=row_num, column=15, value=status_name).alignment = data_alignment
            ws.cell(row=row_num, column=15).fill = row_fill
            ws.cell(row=row_num, column=16, value=violations_str).alignment = data_alignment
            ws.cell(row=row_num, column=16).fill = row_fill
            ws.cell(row=row_num, column=17, value=members_str).alignment = data_alignment
            ws.cell(row=row_num, column=17).fill = row_fill
            
            row_num += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[col_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header row (freeze after statistics section)
    ws.freeze_panes = f"A{header_row + 1}"
    
    # Enable AutoFilter for easy filtering by group, profession, etc.
    # AutoFilter should start from header row
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row_num - 1}"
    
    # Make header row bold and set row height
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 25
    
    # Save file
    filename = f"hesabat_{date}.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    
    return filepath


# ================== HANDLERLƏR ==================

@dp.message(CommandStart())
async def handle_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    full_name = user.full_name if user else "Istifadeci"

    # Ensure all tables exist (legacy + GPS)
    db.init_db()
    db.init_gps_tables()
    db.init_group_codes()
    db.init_registrations()

    # Register minimal user2 for GPS flow
    db.get_or_create_user2(telegram_id=user.id, full_name=full_name)  # type: ignore[arg-type]

    # If admin, show admin menu directly
    if user and is_admin(user.id):
        await message.answer("Xoş gəldiniz!\nAdmin menyusundan istifadə edin:", reply_markup=admin_keyboard())
        return

    # Check legacy user profile; if missing, start registration FSM
    prof = db.get_user_by_telegram_id(user.id) if user else None
    if not prof:
        await state.clear()
        await state.set_state(Reg.profession)
        await message.answer(
            "Qeydiyyat: əvvəlcə peşəni seçin",
            reply_markup=professions_keyboard(),
        )
        return

    # Check if last registration was more than 4 months ago - auto reset
    if prof and isinstance(prof.get("id"), int):
        user_id = int(prof["id"])
        last_reg_date = db.get_last_registration_date(user_id)
        
        if last_reg_date:
            try:
                last_reg = datetime.strptime(last_reg_date, "%Y-%m-%d").date()
                today = now_baku().date()
                days_passed = (today - last_reg).days
                
                # If more than 4 months (approximately 120 days) passed, reset user
                if days_passed > 120:
                    db.delete_user_all(user.id)
                    await state.clear()
                    await state.set_state(Reg.profession)
                    await message.answer(
                        "ℹ️ Son qeydiyyatınızdan 4 aydan çox keçib. Yenidən qeydiyyatdan keçin.\n\nQeydiyyat: əvvəlcə peşəni seçin",
                        reply_markup=professions_keyboard(),
                    )
                    return
            except Exception as e:
                print(f"[handle_start] Error checking last registration: {e}")

    # User is registered, show menu immediately
    prof_name = prof.get("name", "İstifadəçi") if prof else "İstifadəçi"
    welcome_text = (
        f"👋 Xoş gəldiniz, {prof_name}!\n\n"
        "📋 Mənü funksiyaları:\n"
        "• 🟢 Giriş - iş yerinə giriş\n"
        "• 🔴 Çıxış - iş yerindən çıxış\n"
        "\n"
        "Mənüdən istifadə edin:"
    )
    await message.answer(
        welcome_text,
        reply_markup=worker_keyboard(),
    )


@dp.message(Reg.profession)
async def reg_pick_profession(message: Message, state: FSMContext) -> None:
    try:
        text = (message.text or "").strip()
        if text == "❌ Ləğv et":
            await state.clear()
            await message.answer("Ləğv edildi.", reply_markup=worker_keyboard())
            return

        chosen = None
        raw = text.strip('"\' ').strip()

        # 1) numeric prefix like "1." or just number
        idx_part = raw.split(".", 1)[0]
        if idx_part.isdigit():
            idx = int(idx_part) - 1
            if 0 <= idx < len(PROFESSIONS):
                chosen = PROFESSIONS[idx]

        # 2) exact case-insensitive
        if not chosen:
            for p in PROFESSIONS:
                if p.lower() == raw.lower():
                    chosen = p
                    break

        # 3) loose contains match
        if not chosen:
            for p in PROFESSIONS:
                if raw.lower() in p.lower():
                    chosen = p
                    break

        if not chosen:
            await message.answer("Peşə düzgün seçilmədi, siyahıdan seçin.", reply_markup=professions_keyboard())
            return

        await state.update_data(profession=chosen)
        await state.set_state(Reg.code)
        today = today_baku()
        await message.answer(
            f"Peşə: {chosen}\nİndi isə bu günün kodunu daxil edin ({today})",
            reply_markup=ReplyKeyboardRemove(),
        )
    except Exception as e:
        print(f"[reg_pick_profession] error: {e}")
        await state.clear()
        await message.answer("❌ Xəta baş verdi. Qeydiyyatı yenidən başlayın: /start")


@dp.message(Reg.code)
async def reg_enter_code(message: Message, state: FSMContext) -> None:
    try:
        code = (message.text or "").strip()
        data = await state.get_data()
        prof = data.get("profession")
        if not prof:
            await state.set_state(Reg.profession)
            await message.answer("Əvvəl peşə seçin.", reply_markup=professions_keyboard())
            return
        db.init_group_codes()
        if not db.is_group_code_valid(profession=prof, code=code):
            await message.answer("❌ Kod yanlışdır. Yenidən cəhd edin.")
            return
        await state.update_data(code=code)
        await state.set_state(Reg.name)
        await message.answer("Ad Soyad daxil edin:")
    except Exception as e:
        print(f"[reg_enter_code] error: {e}")
        await state.clear()
        await message.answer("❌ Xəta baş verdi. Qeydiyyatı yenidən başlayın: /start")


@dp.message(Reg.name)
async def reg_enter_name(message: Message, state: FSMContext) -> None:
    name = (message.text or "").strip()
    if len(name) < 3:
        await message.answer("Ad ən azı 3 simvol olmalıdır.")
        return
    await state.update_data(name=name)
    await state.set_state(Reg.fin)
    await message.answer("FIN kodunu daxil edin:")


@dp.message(Reg.fin)
async def reg_enter_fin(message: Message, state: FSMContext) -> None:
    fin = (message.text or "").strip().upper()
    if not (7 <= len(fin) <= 10):
        await message.answer("FIN düzgün deyil. Yenidən daxil edin.")
        return
    await state.update_data(fin=fin)
    await state.set_state(Reg.document_series_number)
    await message.answer("Vəsiqənin seriya və nömrəsini daxil edin:")


@dp.message(Reg.document_series_number)
async def reg_enter_document_series_number(message: Message, state: FSMContext) -> None:
    raw_value = (message.text or "").strip()

    def validate_seriya(val: str) -> tuple[bool, str]:
        s = val.replace(" ", "").upper()
        # Nümunələr: AA1234567 (2 hərf + 7 rəqəm), AZE12345678 (passport)
        patterns = [
            r"^[A-Z]{2}\d{7}$",
            r"^[A-Z]{3}\d{8}$",
        ]
        for p in patterns:
            if re.match(p, s):
                return True, s
        return False, (
            "Vəsiqə seriyası/nömrəsi düzgün deyil. Nümunələr: AA1234567 və ya AZE12345678. "
            "Yalnız latın hərfləri və rəqəmlər, boşluq olmadan."
        )

    ok, normalized = validate_seriya(raw_value)
    if not ok:
        await message.answer(normalized)
        return

    await state.update_data(document_series_number=normalized)
    await state.set_state(Reg.phone_number)
    await message.answer("Telefon nömrənizi daxil edin (məs: 501234567 və ya 0501234567):")


def validate_and_normalize_phone(phone: str) -> tuple[bool, str]:
    """
    Telefon nömrəsini yoxlayır və normalize edir.
    Returns: (is_valid, normalized_phone or error_message)
    """
    phone = phone.strip()
    
    # Boş ola bilməz
    if not phone:
        return (False, "Telefon nömrəsi boş ola bilməz. Zəhmət olmasa telefon nömrənizi daxil edin.")
    
    # Yalnız rəqəmlər və + simvolu
    clean_phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")

    # 9 rəqəmli yerli format (məs: 501234567) -> +994501234567
    if len(clean_phone) == 9 and clean_phone.isdigit():
        if clean_phone.startswith(("50", "51", "55", "60", "70", "77", "10", "12", "90", "99")):
            return (True, f"+994{clean_phone}")
        return (False, "Telefon nömrəsi düzgün deyil. Nümunə: 501234567 və ya 0501234567")
    
    # +994 formatında
    if clean_phone.startswith("+994"):
        digits = clean_phone[4:]  # +994-dən sonra
        if len(digits) == 9 and digits.isdigit():
            # Operator kodu ilə başlamalıdır (50, 51, 55, 60, 70, 77, 10, 12, 90, 99)
            if digits.startswith(("50", "51", "55", "60", "70", "77", "10", "12", "90", "99")):
                return (True, f"+994{digits}")
        return (False, "Telefon nömrəsi düzgün deyil. Nümunə: 501234567 və ya 0501234567")
    
    # 994 formatında (+ olmadan)
    if clean_phone.startswith("994"):
        digits = clean_phone[3:]  # 994-dən sonra
        if len(digits) == 9 and digits.isdigit():
            if digits.startswith(("50", "51", "55", "60", "70", "77", "10", "12", "90", "99")):
                return (True, f"+994{digits}")
        return (False, "Telefon nömrəsi düzgün deyil. Nümunə: 501234567 və ya 0501234567")
    
    # 0 ilə başlayır (yerli format: 050, 051, və s.)
    if clean_phone.startswith("0"):
        if len(clean_phone) == 10 and clean_phone.isdigit():
            operator = clean_phone[1:3]  # 0-dan sonra 2 rəqəm
            if operator in ("50", "51", "55", "60", "70", "77", "10", "12"):
                # 0-ı çıxarıb +994 əlavə et
                return (True, f"+994{clean_phone[1:]}")
        return (False, "Telefon nömrəsi düzgün deyil. Nümunə: 501234567 və ya 0501234567")
    
    return (False, "Telefon nömrəsi düzgün formatda deyil. Nümunə: 501234567 və ya 0501234567")


@dp.message(Reg.phone_number)
async def reg_enter_phone_number(message: Message, state: FSMContext) -> None:
    phone_raw = (message.text or "").strip()
    
    # Telefon nömrəsini yoxla və normalize et
    is_valid, result = validate_and_normalize_phone(phone_raw)
    if not is_valid:
        await message.answer(result)
        return
    
    phone_number = result  # Normalize edilmiş telefon nömrəsi
    
    data = await state.get_data()
    code = data.get("code")
    profession = data.get("profession")
    name = data.get("name")
    fin = data.get("fin")
    document_series_number = data.get("document_series_number")
    user = message.from_user
    if not user or not code or not name or not profession or not fin or not document_series_number:
        await state.clear()
        await message.answer("❌ Xəta. Qeydiyyatı yenidən başlayın: /start")
        return

    # Save/Update user profile with all fields
    db.upsert_user_profile(telegram_id=user.id, name=name, fin=fin, code=code, seriya=document_series_number, phone_number=phone_number)

    # Fetch user row to get users.id
    prof = db.get_user_by_telegram_id(user.id)
    today = today_baku()

    # Duplicate protection for same user + profession + code + date
    if prof and isinstance(prof.get("id"), int):
        legacy_user_id = int(prof["id"])  # type: ignore[index]
        if db.has_registration(legacy_user_id, today, profession, code):
            await state.clear()
            await message.answer("ℹ️ Bu gün üçün artıq qeydiyyatınız var.", reply_markup=worker_keyboard())
            return
        db.add_registration(legacy_user_id, today, profession, code)

    await state.clear()
    
    # Təsdiq mesajı
    confirmation_message = (
        "✅ Qeydiyyat tamamlandı!\n\n"
        f"👤 Ad: {name}\n"
        f"🆔 FIN: {fin}\n"
        f"📞 Telefon: {phone_number}\n"
        f"📋 Kod: {code}\n"
        f"💼 Peşə: {profession}\n\n"
        "Menyudan istifadə edin."
    )
    await message.answer(confirmation_message, reply_markup=worker_keyboard())
    
    # Xüsusi xəbərdarlıq mesajı
    warning_message = (
        "⚠️ XƏBƏRDARLIQ\n\n"
        "Giriş-çıxışı yalnız çalışacağınız məkanda vurun. "
        "Əks halda qeydə alınmayacaq.\n\n"
        "📍 GPS aktivləşdirməsi:\n"
        "• Telefonunuzun parametrlərində yerləşmə xidmətlərini açın\n"
        "• Telegram-a lokasiya icazəsi verin\n\n"
        "Bu, giriş və çıxış zamanı lokasiyanızı düzgün qeydə almaq üçün lazımdır."
    )
    await message.answer(warning_message)
    
    # Admin-ə bildiriş göndər
    if ADMIN_ID != 0:
        try:
            await notifications.notify_registration_complete(
                bot=bot,
                admin_id=ADMIN_ID,
                user_name=name,
                user_phone=phone_number,
                user_fin=fin,
                code=code
            )
        except Exception as e:
            print(f"[reg_enter_phone_number] Admin bildirişi xətası: {e}")


# ================== ADMIN ƏMRLƏRİ ==================

@dp.message(Command("bugun"))
async def cmd_bugun(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return

    db.init_registrations()
    today = today_baku()
    rows = db.get_registrations_summary(today)
    if not rows:
        await message.answer("Bu gün üçün qeydiyyat yoxdur.")
        return

    lines: list[str] = []
    for r in rows:
        prof = r.get("profession") or "-"
        code = r.get("code") or "-"
        cnt = r.get("cnt") or 0
        lines.append(f"{prof} | {code} | {cnt}")

    txt = "\n".join(lines)
    for part in chunk_send(txt):
        await message.answer(part)


@dp.message(F.text == "👥 İşçilər")
async def btn_isciler(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    db.init_db()
    workers = db.get_all_workers_status()
    today = today_baku()
    db.init_group_codes()
    active_rows = db.get_group_codes(active_on=today, only_active=True)
    active_codes = {r.get('code') for r in active_rows}
    if active_codes:
        workers = [w for w in workers if (w.get('code') in active_codes)]
    if not workers:
        await message.answer("❌ Heç bir işçi qeydiyyatdan keçməyib.")
        return

    # Group by code
    by_code: dict[str, list[dict]] = {}
    for w in workers:
        code = w.get('code') or '-'
        by_code.setdefault(code, []).append(w)

    msg_lines: list[str] = ["👥 İşçilər siyahısı\n"]
    for code, lst in sorted(by_code.items()):
        msg_lines.append(f"📋 Kod: {code} ({len(lst)})")
        for w in lst:
            msg_lines.append(f"• {w.get('name','?')} | FIN: {w.get('fin','-')}")
        msg_lines.append("")

    text = "\n".join(msg_lines)
    for part in chunk_send(text):
        await message.answer(part)


@dp.message(Command("isciler"))
async def cmd_isciler(message: Message) -> None:
    """Command wrapper so admin sees 'isciler' in left commands menu."""
    await btn_isciler(message)


@dp.message(F.text == "📊 Bu gün")
async def btn_bugun(message: Message) -> None:
    await cmd_bugun(message)


@dp.message(F.text == "🎛 Menyu")
async def btn_menu(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    
    # Get statistics
    stats = db.get_total_registered_students()
    today = today_baku()
    active_today = db.get_active_students_count(today)
    
    lines = [
        "🎛 Admin menyu:",
        "",
        "📊 Statistika:",
        f"• Ümumi qeydiyyatdan keçən: {stats['total']}",
        f"• Aktiv tələbələr: {stats['active']}",
        f"• Kursdan çıxarılan: {stats['inactive']}",
        f"• Bu günün aktiv sayı: {active_today}",
        "",
        "📋 Funksiyalar:",
        "➕ Bugünün kodu — peşə seçib kod əlavə edin",
        "📜 Kodlar — bu günün kodları",
        "📈 Kod üzrə hesabat — tarix+kod ilə siyahı",
        "📥 Excel hesabat — gündəlik/həftəlik/aylıq/kod üzrə",
        "👨‍👩‍👧‍👦 Qruplar — qrup idarəetməsi",
        "🎓 Tələbələr — tələbə idarəetməsi",
    ]
    await message.answer("\n".join(lines), reply_markup=admin_keyboard())


@dp.message(F.text == "➕ Bugünün kodu")
async def btn_add_today_code(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    await state.clear()
    await state.set_state(AdminAddG.profession)
    await message.answer("Peşə seçin:", reply_markup=professions_keyboard())


@dp.message(AdminAddG.profession)
async def adminadd_pick_prof(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return

    chosen = None
    raw = text.strip('"\' ').strip()

    idx_part = raw.split(".", 1)[0]
    if idx_part.isdigit():
        idx = int(idx_part) - 1
        if 0 <= idx < len(PROFESSIONS):
            chosen = PROFESSIONS[idx]

    if not chosen:
        for s in PROFESSIONS:
            if s.lower() == raw.lower():
                chosen = s
                break

    if not chosen:
        for s in PROFESSIONS:
            if raw.lower() in s.lower():
                chosen = s
                break

    if not chosen:
        await message.answer("Düzgün peşə seçin.", reply_markup=professions_keyboard())
        return

    await state.update_data(profession=chosen)
    await state.set_state(AdminAddG.code)
    await message.answer("Bugünün kodunu daxil edin:", reply_markup=ReplyKeyboardRemove())


@dp.message(AdminAddG.code)
async def adminadd_enter_code(message: Message, state: FSMContext) -> None:
    code = (message.text or "").strip()
    if not code:
        await message.answer("Kod boş ola bilməz.")
        return
    data = await state.get_data()
    profession = data.get("profession")
    date = data.get("date")  # Check if date is in state (from AdminManageGroup)
    if not date:
        today = today_baku()
    else:
        today = date
    db.init_group_codes()

    ok = False
    for attempt in range(3):
        try:
            ok = db.add_group_code(profession=profession, date=today, code=str(code), is_active=1)
            break
        except sqlite3.OperationalError as e:
            if 'locked' in str(e).lower():
                await asyncio.sleep(0.5 * (attempt + 1))
                continue
            raise
    await state.clear()
    await message.answer("✅ Yadda saxlandı" if ok else "❌ Xəta", reply_markup=admin_keyboard())


@dp.message(F.text == "📜 Kodlar")
async def btn_view_codes(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    today = today_baku()
    db.init_group_codes()
    rows = db.get_group_codes(date=today, only_active=None)
    if not rows:
        await message.answer("Bu gün üçün kod yoxdur.")
        return
    lines = ["Bu günün kodları:"]
    for r in rows:
        lines.append(f"• {r.get('profession')} → {r.get('code')}")
    await message.answer("\n".join(lines))


@dp.message(F.text == "🗒 Qeydiyyatlar")
async def btn_regs_today(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    today = today_baku()
    rows = db.get_registrations(date=today)
    if not rows:
        await message.answer("Bu gün qeydiyyat yoxdur.")
        return
    lines = ["Bu günün qeydiyyatları:"]
    for r in rows:
        lines.append(f"• {r.get('profession')} | {r.get('code')} — {r.get('name')} (FIN: {r.get('fin')})")
    await message.answer("\n".join(lines))


@dp.message(F.text == "👨‍👩‍👧‍👦 Qruplar")
async def btn_manage_groups(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    
    await state.clear()
    await state.set_state(AdminManageGroup.action)
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Qrup kodu əlavə et")],
            [KeyboardButton(text="🗑 Qrup kodu sil")],
            [KeyboardButton(text="📋 Qrup kodlarını göstər")],
            [KeyboardButton(text="❌ Ləğv et")],
        ],
        resize_keyboard=True
    )
    await message.answer(
        "Qrup idarəetməsi:\n\n"
        "• Qrup kodu əlavə et - yeni qrup kodu əlavə edin\n"
        "• Qrup kodu sil - mövcud qrup kodunu silin\n"
        "• Qrup kodlarını göstər - mövcud kodları görün",
        reply_markup=kb
    )


@dp.message(F.text == "🎓 Tələbələr")
async def btn_manage_students(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    
    await state.clear()
    await state.set_state(AdminManageStudent.action)
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📋 Tələbələri göstər"), KeyboardButton(text="🗑 Tələbə sil")],
            [KeyboardButton(text="✏️ Tələbənin məlumatını dəyiş")],
            [KeyboardButton(text="🔒 Tələbəni deaktiv et"), KeyboardButton(text="🔓 Tələbəni aktiv et")],
            [KeyboardButton(text="🔒 Qrup tələbələrini deaktiv et"), KeyboardButton(text="🔓 Qrup tələbələrini aktiv et")],
            [KeyboardButton(text="❌ Ləğv et")],
        ],
        resize_keyboard=True
    )
    await message.answer(
        "Tələbə idarəetməsi:\n\n"
        "• Tələbələri göstər - tələbələrin siyahısını görün\n"
        "• Tələbə sil - tələbəni sistemdən silin\n"
        "• Tələbəni deaktiv et - tələbə üçün giriş-çıxışı bağlayın\n"
        "• Tələbəni aktiv et - tələbə üçün giriş-çıxışı açın\n"
        "• Qrup tələbələrini deaktiv et - qrupun bütün tələbələrini deaktiv edin\n"
        "• Qrup tələbələrini aktiv et - qrupun bütün tələbələrini aktiv edin",
        reply_markup=kb
    )


@dp.message(F.text == "📡 Loglar")
async def btn_logs_today(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    today = today_baku()
    rows = db.get_attendance_logs(date=today)
    if not rows:
        await message.answer("Bu gün üçün log yoxdur.")
        return
    lines: list[str] = ["Bu günün giriş/çıxış logları:"]
    for r in rows[:30]:  # qısa baxış
        lines.append(
            "\n".join([
                f"• {r.get('profession','-')} | {r.get('code','-')}",
                f"  {r.get('name','?')} (FIN: {r.get('fin','-')})",
                f"  🟢 {r.get('giris_time','-')}  📍 {r.get('giris_loc','-')}",
                f"  🔴 {r.get('cixis_time','-')}  📍 {r.get('cixis_loc','-')}",
            ])
        )
    for part in chunk_send("\n\n".join(lines)):
        await message.answer(part)


@dp.message(F.text == "📈 Kod üzrə hesabat")
async def btn_report_code(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    
    await state.clear()
    await state.set_state(AdminReportByCode.date)
    today = today_baku()
    await message.answer(
        f"Tarixi daxil edin (YYYY-MM-DD). Məs: {today}",
        reply_markup=ReplyKeyboardRemove(),
    )


@dp.message(AdminReportByCode.date)
async def adminreport_date(message: Message, state: FSMContext) -> None:
    date_raw = (message.text or "").strip()
    date = _parse_date_or_today(date_raw)
    if not date:
        await message.answer("Tarix formatı düzgündürmü? YYYY-MM-DD və ya 'bugun' yazın.")
        return
    await state.update_data(date=date)
    await state.set_state(AdminReportByCode.code)
    await message.answer("Kodu daxil edin (məs: 101):")


@dp.message(AdminReportByCode.code)
async def adminreport_code(message: Message, state: FSMContext) -> None:
    code = (message.text or "").strip()
    data = await state.get_data()
    date = data.get("date")
    if not date or not code:
        await state.clear()
        await message.answer("❌ Xəta. Yenidən başlayın: '📈 Kod üzrə hesabat'", reply_markup=admin_keyboard())
        return
    rows = db.get_attendance_logs(date=date, code=code)
    await state.clear()
    if not rows:
        await message.answer("Məlumat tapılmadı.", reply_markup=admin_keyboard())
        return
    lines: list[str] = [f"Hesabat — {date} | Kod: {code}"]
    for r in rows:
        lines.append(
            "\n".join([
                f"• {r.get('name','?')} (FIN: {r.get('fin','-')})",
                f"  Peşə: {r.get('profession','-')}",
                f"  🟢 {r.get('giris_time','-')}  📍 {r.get('giris_loc','-')}",
                f"  🔴 {r.get('cixis_time','-')}  📍 {r.get('cixis_loc','-')}",
            ])
        )
    txt = "\n".join(lines)
    for part in chunk_send(txt):
        await message.answer(part)


@dp.message(F.text == "📥 Excel hesabat")
async def btn_excel_report(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əməliyyat yalnız admin üçündür.")
        return
    
    await state.clear()
    await state.set_state(AdminPeriodReport.period_type)
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Gündəlik"), KeyboardButton(text="📅 Həftəlik")],
            [KeyboardButton(text="📆 Aylıq"), KeyboardButton(text="🔖 Kod üzrə")],
            [KeyboardButton(text="🗓 Tarix aralığı")],
            [KeyboardButton(text="❌ Ləğv et")],
        ],
        resize_keyboard=True
    )
    await message.answer(
        "Hesabat növünü seçin:",
        reply_markup=kb
    )


# ================== HESABAT SİSTEMİ ==================

@dp.message(AdminPeriodReport.period_type)
async def admin_period_type(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return
    
    period_map = {
        "📊 Gündəlik": "daily",
        "📅 Həftəlik": "weekly",
        "📆 Aylıq": "monthly",
        "🔖 Kod üzrə": "code",
        "🗓 Tarix aralığı": "range",
    }
    
    period_type = period_map.get(text)
    if not period_type:
        await message.answer("❌ Zəhmət olmasa butonlardan birini seçin.")
        return
    
    await state.update_data(period_type=period_type)
    
    if period_type == "daily":
        await state.set_state(AdminPeriodReport.start_date)
        today = today_baku()
        await message.answer(
            f"📅 Tarixi daxil edin (YYYY-MM-DD). Məs: {today} və ya 'bugun'",
            reply_markup=ReplyKeyboardRemove()
        )
    elif period_type in ["weekly", "monthly", "range"]:
        await state.set_state(AdminPeriodReport.start_date)
        today = today_baku()
        await message.answer(
            f"📅 Başlanğıc tarixi daxil edin (YYYY-MM-DD). Məs: {today}",
            reply_markup=ReplyKeyboardRemove()
        )
    elif period_type == "code":
        await state.set_state(AdminPeriodReport.code)
        await message.answer(
            "🔖 Qrup kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove()
        )


@dp.message(AdminPeriodReport.start_date)
async def admin_period_start_date(message: Message, state: FSMContext) -> None:
    date_raw = (message.text or "").strip()
    date = _parse_date_or_today(date_raw)
    
    if not date:
        await message.answer("❌ Tarix formatı düzgün deyil. YYYY-MM-DD və ya 'bugun' yazın.")
        return
    
    data = await state.get_data()
    period_type = data.get("period_type")
    
    await state.update_data(start_date=date)
    
    if period_type == "daily":
        # For daily, start_date is the date, set end_date same
        await state.update_data(end_date=date)
        await state.set_state(AdminPeriodReport.format_type)
        await message.answer(
            "Format seçin:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="📊 Excel"), KeyboardButton(text="📄 CSV")],
                    [KeyboardButton(text="❌ Ləğv et")],
                ],
                resize_keyboard=True
            )
        )
    else:
        await state.set_state(AdminPeriodReport.end_date)
        today = today_baku()
        await message.answer(
            f"📅 Bitiş tarixini daxil edin (YYYY-MM-DD). Məs: {today}",
            reply_markup=ReplyKeyboardRemove()
        )


@dp.message(AdminPeriodReport.end_date)
async def admin_period_end_date(message: Message, state: FSMContext) -> None:
    date_raw = (message.text or "").strip()
    date = _parse_date_or_today(date_raw)
    
    if not date:
        await message.answer("❌ Tarix formatı düzgün deyil. YYYY-MM-DD və ya 'bugun' yazın.")
        return
    
    data = await state.get_data()
    start_date = data.get("start_date")
    
    if start_date and date < start_date:
        await message.answer("❌ Bitiş tarixi başlanğıc tarixindən əvvəl ola bilməz.")
        return
    
    await state.update_data(end_date=date)
    await state.set_state(AdminPeriodReport.format_type)
    await message.answer(
        "Format seçin:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📊 Excel"), KeyboardButton(text="📄 CSV")],
                [KeyboardButton(text="❌ Ləğv et")],
            ],
            resize_keyboard=True
        )
    )


@dp.message(AdminPeriodReport.code)
async def admin_period_code(message: Message, state: FSMContext) -> None:
    code = (message.text or "").strip()
    if not code:
        await message.answer("❌ Kod boş ola bilməz.")
        return
    
    await state.update_data(code=code)
    await state.set_state(AdminPeriodReport.start_date)
    today = today_baku()
    await message.answer(
        f"📅 Başlanğıc tarixi daxil edin (YYYY-MM-DD). Məs: {today}",
        reply_markup=ReplyKeyboardRemove()
    )


@dp.message(AdminPeriodReport.format_type)
async def admin_period_format(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return
    
    format_map = {
        "📊 Excel": "excel",
        "📄 CSV": "csv"
    }
    
    format_type = format_map.get(text)
    if not format_type:
        await message.answer("❌ Zəhmət olmasa format seçin.")
        return
    
    data = await state.get_data()
    period_type = data.get("period_type")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    code = data.get("code")
    
    await state.clear()
    
    try:
        # Calculate date range based on period type
        if period_type == "weekly":
            # Calculate week start and end
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = start_dt + timedelta(days=6)
            end_date = end_dt.isoformat()
        elif period_type == "monthly":
            # Calculate month start and end
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            # First day of month
            month_start = start_dt.replace(day=1)
            # Last day of month
            if start_dt.month == 12:
                month_end = start_dt.replace(year=start_dt.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = start_dt.replace(month=start_dt.month + 1, day=1) - timedelta(days=1)
            start_date = month_start.isoformat()
            end_date = month_end.isoformat()
        
        def _has_real_day_data(r: dict) -> bool:
            # Any attendance time, legacy location text, or GPS coords means this row is real
            def _nonempty(v: object) -> bool:
                if v is None:
                    return False
                s = str(v).strip()
                return s != "" and s != "-"

            return (
                _nonempty(r.get('giris_time'))
                or _nonempty(r.get('cixis_time'))
                or _nonempty(r.get('giris_loc'))
                or _nonempty(r.get('cixis_loc'))
                or r.get('start_lat') is not None
                or r.get('start_lon') is not None
                or r.get('end_lat') is not None
                or r.get('end_lon') is not None
            )

        # Get report data with explicit per-row date to avoid all rows having start_date
        report_data: list[dict] = []
        cur_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt2 = datetime.strptime(end_date, "%Y-%m-%d").date()
        while cur_dt <= end_dt2:
            d = cur_dt.isoformat()
            day_rows = db.get_daily_report_for_excel(d)
            if code:
                day_rows = [r for r in day_rows if r.get('code') == code]
            # Period/range hesabatlarında boş sətrləri (heç bir data olmayan) çıxart
            day_rows = [r for r in day_rows if _has_real_day_data(r)]
            for r in day_rows:
                r['date'] = d
                report_data.append(r)
            cur_dt += timedelta(days=1)
        
        if not report_data:
            await message.answer("❌ Seçilən dövrdə məlumat tapılmadı.", reply_markup=admin_keyboard())
            return
        
        # Add status and violations to each row
        for row in report_data:
            giris_time = row.get('giris_time')
            cixis_time = row.get('cixis_time')
            start_lat = row.get('start_lat')
            start_lon = row.get('start_lon')
            end_lat = row.get('end_lat')
            end_lon = row.get('end_lon')
            is_active = row.get('is_active', 1)
            
            status, violations = check_rules_violation(
                giris_time, cixis_time,
                float(start_lat) if start_lat is not None else None,
                float(start_lon) if start_lon is not None else None,
                float(end_lat) if end_lat is not None else None,
                float(end_lon) if end_lon is not None else None,
                is_active,
                CHECKIN_DEADLINE_HOUR, CHECKOUT_DEADLINE_HOUR, MIN_WORK_DURATION_HOURS,
                WORKPLACE_LAT, WORKPLACE_LON, WORKPLACE_RADIUS_M, LOCATION_TOLERANCE_M
            )
            row['status'] = get_status_name(status)
            row['violations'] = "; ".join(violations) if violations else "-"

            # Precompute location fields for both Excel and CSV exports
            row['gps_coords'] = '-'
            row['address'] = '-'
            row['maps_link'] = '-'
            try:
                lat = None
                lon = None
                if start_lat is not None and start_lon is not None:
                    lat = float(start_lat)
                    lon = float(start_lon)
                elif end_lat is not None and end_lon is not None:
                    lat = float(end_lat)
                    lon = float(end_lon)

                if lat is not None and lon is not None:
                    row['gps_coords'] = f"{lat}, {lon}"
                    # Sync context: use coordinates (geocoding cache is async-only)
                    row['address'] = row['gps_coords']
                    row['maps_link'] = f"https://maps.google.com/?q={lat},{lon}"
                else:
                    # Fallback to legacy location text fields
                    giris_loc = (row.get('giris_loc') or '').strip()
                    cixis_loc = (row.get('cixis_loc') or '').strip()
                    address = giris_loc or cixis_loc
                    if address:
                        row['address'] = address
                        encoded_address = urllib.parse.quote(address)
                        row['maps_link'] = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"
            except Exception as e:
                print(f"[admin_period_format] Error computing location fields: {e}")
        
        await message.answer(f"📊 Hesabat hazırlanır... (Format: {format_type.upper()})")
        
        if format_type == "excel":
            # For daily reports, use existing function
            if period_type == "daily":
                filepath = generate_daily_excel_report(start_date)
                filename = f"hesabat_{start_date}.xlsx"
            else:
                # For period reports, create Excel with all dates
                filepath = generate_period_excel_report(report_data, start_date, end_date, code)
                period_name = f"{start_date}_to_{end_date}"
                if code:
                    period_name += f"_{code}"
                filename = f"hesabat_{period_name}.xlsx"
            
            document = FSInputFile(filepath, filename=filename)
            period_str = f"{start_date}"
            if end_date != start_date:
                period_str += f" - {end_date}"
            if code:
                period_str += f" (Kod: {code})"
            
            await message.answer_document(
                document,
                caption=f"📥 Hesabat - {period_str}\n\nFormat: Excel"
            )
            
        elif format_type == "csv":
            period_name = f"{start_date}"
            if end_date != start_date:
                period_name += f"_to_{end_date}"
            if code:
                period_name += f"_{code}"
            filename = f"hesabat_{period_name}.csv"
            filepath = generate_csv_report(report_data, filename)
            
            document = FSInputFile(filepath, filename=filename)
            period_str = f"{start_date}"
            if end_date != start_date:
                period_str += f" - {end_date}"
            if code:
                period_str += f" (Kod: {code})"
            
            await message.answer_document(
                document,
                caption=f"📥 Hesabat - {period_str}\n\nFormat: CSV"
            )
        
        # Clean up
        try:
            os.remove(filepath)
        except Exception:
            pass
            
        await message.answer("✅ Hesabat göndərildi.", reply_markup=admin_keyboard())
        
    except Exception as e:
        print(f"[admin_period_format] error: {e}")
        await message.answer(f"❌ Xəta baş verdi: {str(e)}", reply_markup=admin_keyboard())


def generate_period_excel_report(report_data: List[dict], start_date: str, end_date: str, code: Optional[str] = None) -> str:
    """Generate Excel report for a date range. Similar to daily but includes all dates."""
    # Group by date then by code
    by_date_code: dict[str, dict[str, list[dict]]] = {}
    for row in report_data:
        date = row.get('date', start_date)
        code_key = row.get('code') or '-'
        if date not in by_date_code:
            by_date_code[date] = {}
        if code_key not in by_date_code[date]:
            by_date_code[date][code_key] = []
        by_date_code[date][code_key].append(row)
    
    wb = Workbook()
    ws = wb.active
    period_name = f"{start_date} - {end_date}"
    if code:
        period_name += f" ({code})"
    ws.title = f"Hesabat {period_name}"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Format dates
    months_az = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "İyun", 7: "İyul", 8: "Avqust",
        9: "Sentyabr", 10: "Oktyabr", 11: "Noyabr", 12: "Dekabr"
    }
    
    # Statistics
    stats = db.get_total_registered_students()
    active_on_start = db.get_active_students_count(start_date)
    
    # Write statistics
    stats_row = 1
    ws.cell(row=stats_row, column=1, value="📊 STATİSTİKA").font = Font(bold=True, color="FFFFFF", size=14)
    ws.cell(row=stats_row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{stats_row}:F{stats_row}')
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Ümumi qeydiyyatdan keçən tələbələr:").font = Font(bold=True, size=11)
    ws.cell(row=stats_row, column=2, value=stats['total']).font = Font(bold=True, size=11)
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value="Aktiv tələbələr:").font = Font(bold=True, size=11, color="006100")
    ws.cell(row=stats_row, column=2, value=stats['active']).font = Font(bold=True, size=11, color="006100")
    
    stats_row += 1
    ws.cell(row=stats_row, column=1, value=f"Aktiv tələbə sayı ({start_date}):").font = Font(bold=True, size=11)
    ws.cell(row=stats_row, column=2, value=active_on_start).font = Font(bold=True, size=11)
    
    stats_row += 2
    
    # Headers
    headers = [
        "Tarix", "FIN Kodu", "Ad", "Soyad", "Vəsiqə Seriya", "Telefon",
        "Qrup Kodu", "Peşə", "Giriş Saatı", "Çıxış Saatı",
        "GPS Koordinatları", "Lokasiya", "Xəritə Linki", "Status", "Qayda Pozuntuları"
    ]
    
    header_row = stats_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    row_num = header_row + 1
    for date in sorted(by_date_code.keys()):
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d").date()
            formatted_date = f"{date_obj.day} {months_az[date_obj.month]} {date_obj.year}"
        except:
            formatted_date = date
        
        for code_key, members in sorted(by_date_code[date].items()):
            for member in members:
                name_full = member.get('name', '')
                name_parts = name_full.strip().split(maxsplit=1)
                ad = name_parts[0] if name_parts else name_full
                soyad = name_parts[1] if len(name_parts) > 1 else ""
                
                fin = member.get('fin', '')
                seriya = member.get('seriya') or '-'
                phone = member.get('phone_number') or '-'
                profession = member.get('profession', '-')
                # Normalize entry/exit display similar to daily report
                giris_time_raw = member.get('giris_time')
                cixis_time_raw = member.get('cixis_time')
                if giris_time_raw is None or str(giris_time_raw).strip() == '':
                    giris_time = "Giriş yoxdur"
                    cixis_time = "Giriş yoxdur"
                elif cixis_time_raw is None or str(cixis_time_raw).strip() == '':
                    giris_time = str(giris_time_raw).strip()
                    cixis_time = "Yoxdur"
                else:
                    giris_time = str(giris_time_raw).strip()
                    cixis_time = str(cixis_time_raw).strip()
                
                start_lat = member.get('start_lat')
                start_lon = member.get('start_lon')
                end_lat = member.get('end_lat')
                end_lon = member.get('end_lon')

                gps_coords = member.get('gps_coords') or '-'
                address = member.get('address') or '-'
                maps_link = member.get('maps_link') or '-'

                # Compute if not precomputed
                if gps_coords == '-' and address == '-' and maps_link == '-':
                    try:
                        lat = None
                        lon = None
                        if start_lat is not None and start_lon is not None:
                            lat = float(start_lat)
                            lon = float(start_lon)
                        elif end_lat is not None and end_lon is not None:
                            lat = float(end_lat)
                            lon = float(end_lon)
                        if lat is not None and lon is not None:
                            gps_coords = f"{lat}, {lon}"
                            # Sync context: use coordinates (geocoding cache is async-only)
                            address = gps_coords
                            maps_link = f"https://maps.google.com/?q={lat},{lon}"
                        else:
                            giris_loc = (member.get('giris_loc') or '').strip()
                            cixis_loc = (member.get('cixis_loc') or '').strip()
                            addr = giris_loc or cixis_loc
                            if addr:
                                address = addr
                                encoded_address = urllib.parse.quote(addr)
                                maps_link = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"
                    except Exception as e:
                        print(f"[generate_period_excel_report] Error computing location: {e}")
                
                status = member.get('status', '-')
                violations = member.get('violations', '-')
                
                is_active = member.get('is_active', 1)
                status_color = get_status_color("inactive" if is_active == 0 else ("ok" if status == "Qaydalara uyğundur" else "violation"))
                row_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                
                # Write row
                ws.cell(row=row_num, column=1, value=formatted_date).fill = row_fill
                ws.cell(row=row_num, column=2, value=fin).fill = row_fill
                ws.cell(row=row_num, column=3, value=ad).fill = row_fill
                ws.cell(row=row_num, column=4, value=soyad).fill = row_fill
                ws.cell(row=row_num, column=5, value=seriya).fill = row_fill
                ws.cell(row=row_num, column=6, value=phone).fill = row_fill
                ws.cell(row=row_num, column=7, value=code_key).fill = row_fill
                ws.cell(row=row_num, column=8, value=profession).fill = row_fill
                ws.cell(row=row_num, column=9, value=giris_time).fill = row_fill
                ws.cell(row=row_num, column=10, value=cixis_time).fill = row_fill
                ws.cell(row=row_num, column=11, value=gps_coords).fill = row_fill
                ws.cell(row=row_num, column=12, value=address).fill = row_fill
                if maps_link != '-':
                    cell = ws.cell(row=row_num, column=13, value="Xəritədə bax")
                    cell.hyperlink = maps_link
                    cell.font = Font(color="0000FF", underline="single")
                    cell.fill = row_fill
                else:
                    ws.cell(row=row_num, column=13, value="-").fill = row_fill
                ws.cell(row=row_num, column=14, value=status).fill = row_fill
                ws.cell(row=row_num, column=15, value=violations).fill = row_fill
                
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).alignment = data_alignment
                
                row_num += 1
    
    # Auto-adjust columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws[col_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row_num - 1}"
    
    filename = f"hesabat_{start_date}_to_{end_date}"
    if code:
        filename += f"_{code}"
    filename += ".xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    
    return filepath


@dp.message(Command("excel"))
async def cmd_excel(message: Message) -> None:
    """Command to export today's report to Excel"""
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    
    # Parse date from command if provided
    parts = shlex.split(message.text or "")
    if len(parts) >= 2:
        date = _parse_date_or_today(parts[1])
        if not date:
            await message.answer("❌ Tarix formatı düzgün deyil. YYYY-MM-DD və ya 'bugun' yazın.")
            return
    else:
        date = today_baku()
    
    try:
        await message.answer(f"📊 {date} üçün Excel hesabatı hazırlanır...")
        filepath = generate_daily_excel_report(date)
        
        # Send Excel file
        document = FSInputFile(filepath, filename=f"hesabat_{date}.xlsx")
        await message.answer_document(
            document,
            caption=f"📥 Günlük hesabat - {date}\n\nBütün işçilərin məlumatları, giriş-çıxış saatları və qrup üzvləri daxildir."
        )
        
        # Clean up file after sending
        try:
            os.remove(filepath)
        except Exception:
            pass
            
    except Exception as e:
        print(f"[cmd_excel] error: {e}")
        await message.answer(f"❌ Xəta baş verdi: {str(e)}")


# ================== İSTİFADƏÇİ KOMANDALARI ==================
@dp.message(Command("editprofile"))
async def cmd_editprofile(message: Message, state: FSMContext) -> None:
    """Edit user profile (name, FIN, etc.)"""
    user = message.from_user
    if not user:
        return

    if not is_admin(user.id):
        await state.clear()
        await message.answer(
            "❌ Məlumat dəyişdirmə yalnız admin tərəfindən edilir. Zəhmət olmasa adminə müraciət edin.",
            reply_markup=worker_keyboard(),
        )
        if ADMIN_ID != 0:
            try:
                prof = db.get_user_by_telegram_id(user.id) or {}
                await bot.send_message(
                    ADMIN_ID,
                    "✏️ Məlumat dəyişikliyi sorğusu\n\n"
                    f"👤 İstifadəçi: {prof.get('name', user.full_name or '-') }\n"
                    f"🆔 FIN: {prof.get('fin', '-') }\n"
                    f"📋 Kod: {prof.get('code', '-') }\n"
                    f"🧾 Telegram ID: {user.id}\n\n"
                    "İstifadəçi məlumatını dəyişmək istəyir. Dəyişiklik admin panelindən edilməlidir.",
                )
            except Exception:
                pass
        return
    
    prof = db.get_user_by_telegram_id(user.id)
    if not prof:
        await message.answer("❌ Qeydiyyatınız tapılmadı. Əvvəlcə /start ilə qeydiyyatdan keçin.")
        return
    
    await state.clear()
    await state.set_state(EditProfile.field)
    
    current_info = (
        f"📋 Hazırkı məlumatlarınız:\n\n"
        f"👤 Ad: {prof.get('name', '-')}\n"
        f"🆔 FIN: {prof.get('fin', '-')}\n"
        f"📋 Kod: {prof.get('code', '-')}\n\n"
        f"Hansı məlumatı dəyişdirmək istəyirsiniz?"
    )
    
    edit_kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👤 Adı dəyişdir"), KeyboardButton(text="🆔 FIN-i dəyişdir")],
            [KeyboardButton(text="❌ Ləğv et")],
        ],
        resize_keyboard=True
    )
    
    await message.answer(current_info, reply_markup=edit_kb)


@dp.message(EditProfile.field)
async def editprofile_field(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "❌ Ləğv et" or text.lower() in ("ləğv", "cancel", "legv"):
        await state.clear()
        await message.answer("❌ Ləğv edildi.", reply_markup=worker_keyboard())
        return
    
    field_map = {
        "👤 adı dəyişdir": "name",
        "adı dəyişdir": "name",
        "ad": "name",
        "🆔 fin-i dəyişdir": "fin",
        "fin-i dəyişdir": "fin",
        "fin": "fin",
    }
    
    field = field_map.get(text.lower())
    if not field:
        await message.answer("❌ Zəhmət olmasa butonlardan birini seçin.")
        return
    
    await state.update_data(field=field)
    await state.set_state(EditProfile.new_value)
    
    field_name = "Ad" if field == "name" else "FIN"
    await message.answer(
        f"Yeni {field_name} daxil edin:",
        reply_markup=ReplyKeyboardRemove()
    )


@dp.message(EditProfile.new_value)
async def editprofile_value(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if not user:
        return
    
    new_value = (message.text or "").strip()
    if not new_value:
        await message.answer("❌ Boş dəyər daxil edilə bilməz.")
        return
    
    data = await state.get_data()
    field = data.get("field")
    
    if not field:
        await state.clear()
        await message.answer("❌ Xəta. Yenidən başlayın: /editprofile")
        return
    
    prof = db.get_user_by_telegram_id(user.id)
    if not prof:
        await state.clear()
        await message.answer("❌ Qeydiyyatınız tapılmadı.")
        return
    
    # Update profile
    if field == "name":
        db.upsert_user_profile(
            telegram_id=user.id,
            name=new_value,
            fin=prof.get("fin", ""),
            code=prof.get("code", ""),
            seriya=prof.get("seriya", ""),
            phone_number=prof.get("phone_number", "")
        )
        await message.answer(f"✅ Ad dəyişdirildi: {new_value}", reply_markup=worker_keyboard())
    elif field == "fin":
        if not (7 <= len(new_value) <= 10):
            await message.answer("❌ FIN düzgün deyil. 7-10 simvol olmalıdır.")
            return
        db.upsert_user_profile(
            telegram_id=user.id,
            name=prof.get("name", ""),
            fin=new_value.upper(),
            code=prof.get("code", ""),
            seriya=prof.get("seriya", ""),
            phone_number=prof.get("phone_number", "")
        )
        await message.answer(f"✅ FIN dəyişdirildi: {new_value.upper()}", reply_markup=worker_keyboard())
    
    await state.clear()


@dp.message(Command("professions"))
async def cmd_professions(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    txt = "Peşələr:\n" + "\n".join(f"{i+1}. {p}" for i, p in enumerate(PROFESSIONS))
    await message.answer(txt)


@dp.message(Command("addgcode"))
async def cmd_addgcode(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    db.init_group_codes()
    parts = shlex.split(message.text or "")
    if len(parts) < 4:
        await message.answer("İstifadə: /addgcode \"Peşə\" YYYY-MM-DD KOD [1|0]")
        return
    prof = parts[1]
    date = parts[2]
    code = parts[3]
    is_active = 1
    if len(parts) >= 5 and parts[4] in ("0", "1"):
        is_active = int(parts[4])
    if prof.isdigit():
        idx = int(prof) - 1
        if 0 <= idx < len(PROFESSIONS):
            prof = PROFESSIONS[idx]
    else:
        m = _match_profession(prof)
        if m:
            prof = m
    if prof not in PROFESSIONS:
        await message.answer("Peşə tapılmadı. /professions ilə siyahıya baxın.")
        return

    ok = db.add_group_code(profession=prof, date=date, code=code, is_active=is_active)
    await message.answer("✅ Yadda saxlandı" if ok else "❌ Xəta")


@dp.message(Command("listgcodes"))
async def cmd_listgcodes(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    db.init_group_codes()
    parts = shlex.split(message.text or "")
    date = None
    only_active = None
    if len(parts) >= 2:
        if parts[1] in ("0", "1"):
            only_active = True if parts[1] == "1" else False
        else:
            date = parts[1]
    if len(parts) >= 3:
        if parts[2] in ("0", "1"):
            only_active = True if parts[2] == "1" else False
    rows = db.get_group_codes(date=date, only_active=only_active)
    if not rows:
        await message.answer("Məlumat tapılmadı.")
        return
    lines: list[str] = ["Qrup kodları:"]
    for r in rows:
        lines.append(f"• {r.get('date')} | {r.get('profession')} → {r.get('code')}")
    txt = "\n".join(lines)
    for part in chunk_send(txt):
        await message.answer(part)


@dp.message(Command("listregs"))
async def cmd_listregs(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    db.init_registrations()
    parts = shlex.split(message.text or "")
    date = None
    profession = None
    code = None
    if len(parts) >= 2:
        date = parts[1]
    if len(parts) >= 3:
        p = parts[2]
        if p.isdigit():
            idx = int(p) - 1
            if 0 <= idx < len(PROFESSIONS):
                profession = PROFESSIONS[idx]
        else:
            m = _match_profession(p)
            if m:
                profession = m
    if len(parts) >= 4:
        code = parts[3]
    rows = db.get_registrations(date=date, profession=profession, code=code)
    if not rows:
        await message.answer("Qeydiyyat tapılmadı.")
        return
    lines = ["Qeydiyyatlar:"]
    for r in rows:
        lines.append(f"• {r.get('date')} | {r.get('profession')} | {r.get('code')} — {r.get('name')} (FIN: {r.get('fin')})")
    txt = "\n".join(lines)
    for part in chunk_send(txt):
        await message.answer(part)


# ================== QRUPLAR VƏ TƏLƏBƏLƏR İDARƏETMƏSİ ==================

@dp.message(AdminManageGroup.action)
async def admin_manage_group_action(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return
    
    if text == "➕ Qrup kodu əlavə et":
        await state.update_data(action="add_code")
        await state.set_state(AdminManageGroup.profession)
        await message.answer("Kod əlavə etmək üçün peşə seçin:", reply_markup=professions_keyboard())
    elif text == "🗑 Qrup kodu sil":
        await state.update_data(action="delete_code")
        await state.set_state(AdminManageGroup.profession)
        await message.answer("Silinəcək kod üçün peşə seçin:", reply_markup=professions_keyboard())
    elif text == "📋 Qrup kodlarını göstər":
        await state.clear()
        today = today_baku()
        db.init_group_codes()
        rows = db.get_group_codes(active_on=today, only_active=None)
        if not rows:
            await message.answer("Aktiv kod yoxdur.", reply_markup=admin_keyboard())
            return
        lines = ["Aktiv kodlar:"]
        for r in rows:
            lines.append(f"• {r.get('profession')} → {r.get('code')}")
        await message.answer("\n".join(lines), reply_markup=admin_keyboard())
    else:
        await message.answer("❌ Zəhmət olmasa butonlardan birini seçin.")


@dp.message(AdminManageGroup.profession)
async def admin_manage_group_profession(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return
    
    chosen = None
    raw = text.strip('"\' ').strip()
    idx_part = raw.split(".", 1)[0]
    if idx_part.isdigit():
        idx = int(idx_part) - 1
        if 0 <= idx < len(PROFESSIONS):
            chosen = PROFESSIONS[idx]
    
    if not chosen:
        for s in PROFESSIONS:
            if s.lower() == raw.lower():
                chosen = s
                break
    
    if not chosen:
        for s in PROFESSIONS:
            if raw.lower() in s.lower():
                chosen = s
                break
    
    if not chosen:
        await message.answer("Düzgün peşə seçin.", reply_markup=professions_keyboard())
        return
    
    await state.update_data(profession=chosen)
    await state.set_state(AdminManageGroup.date)
    await message.answer("Tarixi daxil edin (YYYY-MM-DD) və ya 'bugun' yazın:", reply_markup=ReplyKeyboardRemove())


@dp.message(AdminManageGroup.date)
async def admin_manage_group_date(message: Message, state: FSMContext) -> None:
    date_raw = (message.text or "").strip()
    date = _parse_date_or_today(date_raw)
    
    if not date:
        await message.answer("❌ Tarix formatı düzgün deyil. YYYY-MM-DD və ya 'bugun' yazın.")
        return
    
    await state.update_data(date=date)
    data = await state.get_data()
    action = data.get("action")
    profession = data.get("profession")
    
    if action == "add_code":
        # Use AdminAddG.code state but keep our data
        await state.set_state(AdminAddG.code)
        await message.answer("Qrup kodunu daxil edin:", reply_markup=ReplyKeyboardRemove())
    elif action == "delete_code":
        await state.set_state(AdminManageGroup.code)
        await message.answer("Silinəcək qrup kodunu daxil edin:", reply_markup=ReplyKeyboardRemove())
    else:
        await state.clear()
        await message.answer("❌ Xəta.", reply_markup=admin_keyboard())


@dp.message(AdminManageGroup.code)
async def admin_manage_group_code(message: Message, state: FSMContext) -> None:
    code = (message.text or "").strip()
    if not code:
        await message.answer("❌ Kod boş ola bilməz.")
        return

    data = await state.get_data()
    action = data.get("action")
    profession = data.get("profession")
    date = data.get("date")

    if action != "delete_code" or not profession or not date:
        await state.clear()
        await message.answer("❌ Xəta.", reply_markup=admin_keyboard())
        return

    try:
        db.init_group_codes()
        ok = False
        for attempt in range(3):
            try:
                ok = db.delete_group_code(profession=profession, date=date, code=code)
                break
            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower():
                    await asyncio.sleep(0.5 * (attempt + 1))
                    continue
                raise
        await state.clear()
        if ok:
            await message.answer("✅ Qrup kodu silindi.", reply_markup=admin_keyboard())
        else:
            await message.answer("ℹ️ Bu tarix, peşə və kod üçün məlumat tapılmadı.", reply_markup=admin_keyboard())
    except Exception as e:
        await state.clear()
        print(f"[admin_manage_group_code] delete_code error: {e}")
        await message.answer("❌ Xəta baş verdi.", reply_markup=admin_keyboard())


@dp.message(AdminManageStudent.action)
async def admin_manage_student_action(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "❌ Ləğv et":
        await state.clear()
        await message.answer("Ləğv edildi.", reply_markup=admin_keyboard())
        return
    
    if text == "📋 Tələbələri göstər":
        await state.clear()
        try:
            users = db.get_all_users_with_status()
            if not users:
                await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
                return
            
            lines = ["📋 Tələbələr:\n"]
            active_count = 0
            inactive_count = 0
            
            # Qruplar üzrə qruplaşdır
            by_code: dict[str, list[dict]] = {}
            for u in users:
                code = u.get('code') or 'Kod yoxdur'
                if code not in by_code:
                    by_code[code] = []
                by_code[code].append(u)
            
            # Hər qrup üçün göstər
            for code, code_users in sorted(by_code.items()):
                lines.append(f"📁 Qrup: {code} ({len(code_users)} tələbə)")
                for u in code_users[:20]:  # Hər qrupda max 20 nəfər
                    status_icon = "✅" if u.get('is_active', 1) == 1 else "❌"
                    name = u.get('name', '?')
                    fin = u.get('fin', '-')
                    phone = u.get('phone_number', '-')
                    status_text = "Aktiv" if u.get('is_active', 1) == 1 else "Deaktiv"
                    lines.append(f"  {status_icon} {name} (FIN: {fin}, Tel: {phone}) - {status_text}")
                    if u.get('is_active', 1) == 1:
                        active_count += 1
                    else:
                        inactive_count += 1
                if len(code_users) > 20:
                    lines.append(f"  ... və {len(code_users) - 20} tələbə daha")
                lines.append("")
            
            # Ümumi statistika
            lines.append(f"📊 Statistikalar:")
            lines.append(f"• Ümumi: {len(users)} tələbə")
            lines.append(f"• Aktiv: {active_count} tələbə")
            lines.append(f"• Deaktiv: {inactive_count} tələbə")
            
            text = "\n".join(lines)
            for part in chunk_send(text):
                await message.answer(part)
            await message.answer("Menyu:", reply_markup=admin_keyboard())
        except Exception as e:
            print(f"[admin_manage_student_action] Error listing students: {e}")
            await message.answer("❌ Xəta baş verdi. Yenidən yoxlayın.", reply_markup=admin_keyboard())

    elif text == "🗑 Tələbə sil":
        await state.update_data(action="delete")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Silinəcək tələbənin Telegram ID-sini və ya FIN kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif text == "✏️ Tələbənin məlumatını dəyiş":
        await state.update_data(action="edit_profile")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Dəyişiləcək tələbənin Telegram ID-sini və ya FIN kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif text == "🔒 Tələbəni deaktiv et":
        await state.update_data(action="deactivate")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Deaktiv ediləcək tələbənin Telegram ID-sini və ya FIN kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif text == "🔓 Tələbəni aktiv et":
        await state.update_data(action="activate")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Aktiv ediləcək tələbənin Telegram ID-sini və ya FIN kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif text == "🔒 Qrup tələbələrini deaktiv et":
        await state.update_data(action="deactivate_group")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Deaktiv ediləcək qrup kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    elif text == "🔓 Qrup tələbələrini aktiv et":
        await state.update_data(action="activate_group")
        await state.set_state(AdminManageStudent.code_or_id)
        await message.answer(
            "Aktiv ediləcək qrup kodunu daxil edin:",
            reply_markup=ReplyKeyboardRemove(),
        )
    else:
        await message.answer("❌ Zəhmət olmasa butonlardan birini seçin.")


@dp.message(AdminManageStudent.field)
async def admin_manage_student_edit_field(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()

    if text == "❌ Ləğv et" or text.lower() in ("ləğv", "cancel", "legv"):
        await state.clear()
        await message.answer("❌ Ləğv edildi.", reply_markup=admin_keyboard())
        return

    field_map = {
        "👤 adı dəyişdir": "name",
        "adı dəyişdir": "name",
        "ad": "name",
        "🆔 fin-i dəyişdir": "fin",
        "fin-i dəyişdir": "fin",
        "fin": "fin",
    }
    field = field_map.get(text.lower())
    if not field:
        await message.answer("❌ Zəhmət olmasa butonlardan birini seçin.")
        return

    await state.update_data(field=field)
    await state.set_state(AdminManageStudent.new_value)

    field_name = "Ad" if field == "name" else "FIN"
    await message.answer(
        f"Yeni {field_name} daxil edin:",
        reply_markup=ReplyKeyboardRemove(),
    )


@dp.message(AdminManageStudent.new_value)
async def admin_manage_student_edit_value(message: Message, state: FSMContext) -> None:
    new_value = (message.text or "").strip()
    if not new_value:
        await message.answer("❌ Boş dəyər daxil edilə bilməz.")
        return

    data = await state.get_data()
    field = data.get("field")
    target_telegram_id = data.get("user_id")

    if not field or not target_telegram_id:
        await state.clear()
        await message.answer("❌ Xəta. Yenidən başlayın.", reply_markup=admin_keyboard())
        return

    prof = db.get_user_by_telegram_id(int(target_telegram_id))
    if not prof:
        await state.clear()
        await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
        return

    if field == "fin":
        if not (7 <= len(new_value) <= 10):
            await message.answer("❌ FIN düzgün deyil. 7-10 simvol olmalıdır.")
            return
        new_value = new_value.upper()

    if field == "name":
        name = new_value
        fin = prof.get("fin", "")
    else:
        name = prof.get("name", "")
        fin = new_value

    db.upsert_user_profile(
        telegram_id=int(target_telegram_id),
        name=name,
        fin=fin,
        code=prof.get("code", ""),
        seriya=prof.get("seriya", ""),
        phone_number=prof.get("phone_number", "") or "",
    )

    await state.clear()
    changed_field = "Ad" if field == "name" else "FIN"
    await message.answer(f"✅ {changed_field} dəyişdirildi.", reply_markup=admin_keyboard())


@dp.message(AdminManageStudent.code_or_id)
async def admin_manage_student_code_or_id(message: Message, state: FSMContext) -> None:
    code_or_id = (message.text or "").strip()
    if not code_or_id:
        await message.answer("❌ Boş dəyər daxil edilə bilməz.")
        return
    
    data = await state.get_data()
    action = data.get("action")
    
    if action == "delete":
        # Try to find user by telegram_id or fin
        user = None
        if code_or_id.isdigit():
            user = db.get_user_by_telegram_id(int(code_or_id))
        
        if not user:
            # Try by FIN
            users = db.get_all_users_with_status()
            for u in users:
                if u.get('fin', '').upper() == code_or_id.upper():
                    user = u
                    break
        
        if not user:
            await state.clear()
            await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
            return
        
        await state.update_data(user_id=user.get('telegram_id'), user_name=user.get('name'))
        await state.set_state(AdminManageStudent.confirm)
        await message.answer(
            f"⚠️ Təsdiq:\n\n"
            f"Tələbə: {user.get('name')}\n"
            f"FIN: {user.get('fin')}\n"
            f"Kod: {user.get('code')}\n\n"
            f"Bu tələbəni silmək istədiyinizə əminsiniz?",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[
                    [KeyboardButton(text="✅ Bəli, sil"), KeyboardButton(text="❌ Xeyr")],
                ],
                resize_keyboard=True
            )
        )
    elif action == "edit_profile":
        user = None
        if code_or_id.isdigit():
            user = db.get_user_by_telegram_id(int(code_or_id))

        if not user:
            users = db.get_all_users_with_status()
            for u in users:
                if u.get('fin', '').upper() == code_or_id.upper():
                    user = u
                    break

        if not user:
            await state.clear()
            await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
            return

        await state.update_data(user_id=user.get('telegram_id'), user_name=user.get('name'))
        await state.set_state(AdminManageStudent.field)

        current_info = (
            f"📋 Hazırkı məlumatlar:\n\n"
            f"👤 Ad: {user.get('name', '-')}\n"
            f"🆔 FIN: {user.get('fin', '-')}\n"
            f"📋 Kod: {user.get('code', '-')}\n\n"
            f"Hansı məlumatı dəyişdirmək istəyirsiniz?"
        )

        edit_kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="👤 Adı dəyişdir"), KeyboardButton(text="🆔 FIN-i dəyişdir")],
                [KeyboardButton(text="❌ Ləğv et")],
            ],
            resize_keyboard=True
        )

        await message.answer(current_info, reply_markup=edit_kb)
    elif action == "deactivate":
        user = None
        if code_or_id.isdigit():
            user = db.get_user_by_telegram_id(int(code_or_id))
        
        if not user:
            users = db.get_all_users_with_status()
            for u in users:
                if u.get('fin', '').upper() == code_or_id.upper():
                    user = u
                    break
        
        if not user:
            await state.clear()
            await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
            return
        
        ok = db.set_user_active(user.get('telegram_id'), 0)
        await state.clear()
        if ok:
            await message.answer(f"✅ Tələbə deaktiv edildi: {user.get('name')}", reply_markup=admin_keyboard())
        else:
            await message.answer("❌ Xəta baş verdi.", reply_markup=admin_keyboard())
    elif action == "activate":
        user = None
        if code_or_id.isdigit():
            user = db.get_user_by_telegram_id(int(code_or_id))
        
        if not user:
            users = db.get_all_users_with_status()
            for u in users:
                if u.get('fin', '').upper() == code_or_id.upper():
                    user = u
                    break
        
        if not user:
            await state.clear()
            await message.answer("❌ Tələbə tapılmadı.", reply_markup=admin_keyboard())
            return
        
        ok = db.set_user_active(user.get('telegram_id'), 1)
        await state.clear()
        if ok:
            await message.answer(f"✅ Tələbə aktiv edildi: {user.get('name')}", reply_markup=admin_keyboard())
        else:
            await message.answer("❌ Xəta baş verdi.", reply_markup=admin_keyboard())
    elif action == "deactivate_group":
        try:
            count = db.deactivate_user_by_code(code_or_id)
            await state.clear()
            if count > 0:
                await message.answer(f"✅ {count} tələbə deaktiv edildi (Kod: {code_or_id})", reply_markup=admin_keyboard())
            else:
                await message.answer(f"ℹ️ Bu kod üçün tələbə tapılmadı (Kod: {code_or_id})", reply_markup=admin_keyboard())
        except Exception as e:
            print(f"[admin_manage_student_code_or_id] Error deactivating group: {e}")
            await state.clear()
            await message.answer("❌ Xəta baş verdi. Yenidən yoxlayın.", reply_markup=admin_keyboard())
    elif action == "activate_group":
        try:
            # Activate only inactive users in the group
            users = db.get_users_by_code(code_or_id, only_active=False)
            changed = 0
            for u in users:
                if u.get('is_active', 1) == 0 and u.get('telegram_id') is not None:
                    if db.set_user_active(int(u['telegram_id']), 1):
                        changed += 1
            await state.clear()
            if changed > 0:
                await message.answer(f"✅ {changed} tələbə aktiv edildi (Kod: {code_or_id})", reply_markup=admin_keyboard())
            else:
                await message.answer(f"ℹ️ Aktiv ediləcək deaktiv tələbə tapılmadı (Kod: {code_or_id})", reply_markup=admin_keyboard())
        except Exception as e:
            print(f"[admin_manage_student_code_or_id] Error activating group: {e}")
            await state.clear()
            await message.answer("❌ Xəta baş verdi. Yenidən yoxlayın.", reply_markup=admin_keyboard())


@dp.message(AdminManageStudent.confirm)
async def admin_manage_student_confirm(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    
    if text == "✅ Bəli, sil":
        data = await state.get_data()
        user_id = data.get("user_id")
        user_name = data.get("user_name")
        
        if user_id:
            ok = db.delete_user_by_telegram_id(user_id)
            await state.clear()
            if ok:
                await message.answer(f"✅ Tələbə silindi: {user_name}", reply_markup=admin_keyboard())
            else:
                await message.answer("❌ Xəta baş verdi.", reply_markup=admin_keyboard())
        else:
            await state.clear()
            await message.answer("❌ Xəta.", reply_markup=admin_keyboard())
    elif text == "❌ Xeyr" or text == "❌ Xeyr, ləğv et" or text.lower() in ("xeyr", "ləğv", "cancel", "no"):
        await state.clear()
        await message.answer("❌ Silmə əməliyyatı ləğv edildi.", reply_markup=admin_keyboard())
    else:
        await message.answer("❌ Zəhmət olmasa '✅ Bəli, sil' və ya '❌ Xeyr, ləğv et' seçin.")


@dp.message(Command("logs"))
async def cmd_logs(message: Message) -> None:
    user = message.from_user
    if not user or not is_admin(user.id):
        await message.answer("❌ Bu əmr yalnız admin üçündür.")
        return
    parts = shlex.split(message.text or "")
    date = None
    profession = None
    code = None
    if len(parts) >= 2:
        date = parts[1]
    if len(parts) >= 3:
        p = parts[2]
        if p.isdigit():
            idx = int(p) - 1
            if 0 <= idx < len(PROFESSIONS):
                profession = PROFESSIONS[idx]
        else:
            m = _match_profession(p)
            if m:
                profession = m
    if len(parts) >= 4:
        code = parts[3]
    rows = db.get_attendance_logs(date=date, profession=profession, code=code)
    if not rows:
        await message.answer("Log tapılmadı.")
        return
    lines: list[str] = ["Giriş/Çıxış logları:"]
    for r in rows:
        lines.append(
            "\n".join([
                f"• {r.get('date')} | {r.get('profession','-')} | {r.get('code','-')}",
                f"  {r.get('name','?')} (FIN: {r.get('fin','-')})",
                f"  🟢 {r.get('giris_time','-')}  📍 {r.get('giris_loc','-')}",
                f"  🔴 {r.get('cixis_time','-')}  📍 {r.get('cixis_loc','-')}",
            ])
        )
    txt = "\n".join(lines)
    for part in chunk_send(txt):
        await message.answer(part)


# ================== GİRİŞ / ÇIXIŞ ==================

@dp.message(F.text == "🟢 Giriş")
async def handle_giris(message: Message) -> None:
    user = message.from_user
    if not user:
        return
    user_id = user.id

    now = now_baku()
    if now.hour >= CHECKIN_DEADLINE_HOUR:
        await message.answer(
            f"❌ Giriş {CHECKIN_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
            f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
            reply_markup=worker_keyboard(),
        )
        pending_action.pop(user_id, None)
        return

    # Check if user is active
    prof = db.get_user_by_telegram_id(user_id)
    if prof and prof.get("is_active", 1) == 0:
        await message.answer(
            "❌ Sizin giriş-çıxış hüququnuz deaktiv edilib. "
            "Zəhmət olmasa admin ilə əlaqə saxlayın.",
            reply_markup=worker_keyboard()
        )
        return

    try:
        db.init_gps_tables()
        uid = db.get_or_create_user2(telegram_id=user_id, full_name=user.full_name)
        today = today_baku()
        existing = db.get_user_session_on_date(uid, today)
        if existing:
            await message.answer("ℹ️ Bu gün artıq giriş etmisiniz.", reply_markup=worker_keyboard())
            pending_action.pop(user_id, None)
            return
    except Exception as e:
        print(f"[handle_giris] db error: {e}")

    pending_action[user_id] = ("checkin", time.time())
    await message.answer(
        "📍 Giriş üçün lokasiya göndər",
        reply_markup=location_keyboard("Lokasiyanı göndər"),
    )


@dp.message(F.text == "🔴 Çıxış")
async def handle_cixis(message: Message) -> None:
    user = message.from_user
    if not user:
        return
    user_id = user.id

    now = now_baku()
    if now.hour >= CHECKOUT_DEADLINE_HOUR:
        await message.answer(
            f"❌ Çıxış {CHECKOUT_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
            f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
            reply_markup=worker_keyboard(),
        )
        pending_action.pop(user_id, None)
        return

    # Check if user is active
    prof = db.get_user_by_telegram_id(user_id)
    if prof and prof.get("is_active", 1) == 0:
        await message.answer(
            "❌ Sizin giriş-çıxış hüququnuz deaktiv edilib. "
            "Zəhmət olmasa admin ilə əlaqə saxlayın.",
            reply_markup=worker_keyboard()
        )
        return

    try:
        db.init_gps_tables()
        uid = db.get_or_create_user2(telegram_id=user_id, full_name=user.full_name)
        sess = db.get_open_session(uid)
        if not sess:
            await message.answer(
                "❌ Giriş etmədiyiniz üçün çıxış edə bilmirsiniz. Əvvəlcə giriş edin.",
                reply_markup=worker_keyboard(),
            )
            pending_action.pop(user_id, None)
            return
        else:
            # Early safeguard: do not even prompt for location if minimum work duration not passed
            try:
                start_time = parse_dt_to_baku(sess["start_time"])  # type: ignore[index]
                now = now_baku()
                duration_hours = (now - start_time).total_seconds() / 3600.0
                duration_min = max(0, int((now - start_time).total_seconds() // 60))
                if duration_hours < MIN_WORK_DURATION_HOURS:
                    await message.answer(
                        f"❌ Ən azı {MIN_WORK_DURATION_HOURS} saat sonra çıxış edə bilərsiniz.\n\n"
                        f"⏱ Hal-hazırda keçən vaxt: {duration_min} dəqiqə ({duration_hours:.1f} saat)\n"
                        f"📅 Giriş vaxtı: {start_time.strftime('%H:%M')}",
                        reply_markup=worker_keyboard()
                    )
                    return
            except Exception:
                pass
    except Exception as e:
        print(f"[handle_cixis] db error: {e}")

    pending_action[user_id] = ("checkout", time.time())
    await message.answer(
        "📍 Çıxış üçün lokasiya göndər",
        reply_markup=location_keyboard("Lokasiyanı göndər"),
    )


@dp.message(Command("giris"))
async def cmd_giris(message: Message) -> None:
    await handle_giris(message)


@dp.message(Command("cixis"))
async def cmd_cixis(message: Message) -> None:
    await handle_cixis(message)


@dp.message(Command("help"))
async def cmd_help(message: Message) -> None:
    user = message.from_user
    if not user:
        return
    
    help_text = (
        "📚 Kömək\n\n"
        "🟢 Giriş - İş yerinə giriş etmək üçün düyməyə basın və lokasiyanızı göndərin.\n\n"
        "🔴 Çıxış - İş yerindən çıxış etmək üçün düyməyə basın və lokasiyanızı göndərin.\n\n"
        "📌 Əsas qaydalar:\n"
        "• Giriş-çıxışı yalnız iş yerində vurun\n"
        "• GPS aktiv olmalıdır\n"
        "• Giriş 11:00-a qədər olmalıdır\n"
        "• Çıxış 19:00-a qədər olmalıdır\n"
        "• Minimum iş müddəti 6 saatdır\n\n"
        "Suallarınız üçün admin ilə əlaqə saxlayın."
    )
    
    if user and is_admin(user.id):
        help_text += "\n\n🔧 Admin funksiyaları üçün admin menyusundan istifadə edin."
        await message.answer(help_text, reply_markup=admin_keyboard())
    else:
        await message.answer(help_text, reply_markup=worker_keyboard())


# ================== LOKASİYA HANDLER ==================

@dp.message(F.location)
async def handle_location(message: Message) -> None:
    try:
        if not message.location:
            return

        user = message.from_user
        if not user:
            return

        user_id = user.id
        pa = pending_action.get(user_id)
        if not pa:
            # Lokasiya spami üçün cavab vermirik
            return
        action, ts = pa

        # Expire old intents
        if time.time() - ts > LOCATION_TIMEOUT:
            pending_action.pop(user_id, None)
            await message.answer("⏱ Lokasiya çox gec gəldi, yenidən giriş/çıxış seçin.", reply_markup=worker_keyboard())
            return

        db.init_db()
        db.init_gps_tables()

        lat = float(message.location.latitude)
        lon = float(message.location.longitude)

        uid = db.get_or_create_user2(telegram_id=user_id, full_name=user.full_name)

        now = now_baku()
        now_iso = now.isoformat(timespec="seconds")
        today = now.date().isoformat()

        if action == "checkin" and now.hour >= CHECKIN_DEADLINE_HOUR:
            prof = db.get_user_by_telegram_id(user_id)
            name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
            user_phone = prof.get("phone_number") if prof else None

            await message.answer(
                f"❌ Giriş {CHECKIN_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
                f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
                reply_markup=worker_keyboard(),
            )
            pending_action.pop(user_id, None)
            if ADMIN_ID != 0:
                await notifications.notify_rule_violation(
                    bot=bot,
                    admin_id=ADMIN_ID,
                    user_id=user_id,
                    user_name=name,
                    user_phone=user_phone,
                    violation_type="Gecikmə - giriş vaxtı keçib",
                    details=f"Giriş {CHECKIN_DEADLINE_HOUR}:00-dan sonra vurulmağa cəhd edildi. Cari vaxt: {now.strftime('%H:%M')}" ,
                )
            return

        if action == "checkout" and now.hour >= CHECKOUT_DEADLINE_HOUR:
            prof = db.get_user_by_telegram_id(user_id)
            name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
            user_phone = prof.get("phone_number") if prof else None

            await message.answer(
                f"❌ Çıxış {CHECKOUT_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
                f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
                reply_markup=worker_keyboard(),
            )
            pending_action.pop(user_id, None)
            if ADMIN_ID != 0:
                await notifications.notify_rule_violation(
                    bot=bot,
                    admin_id=ADMIN_ID,
                    user_id=user_id,
                    user_name=name,
                    user_phone=user_phone,
                    violation_type="Gecikmə - çıxış vaxtı keçib",
                    details=f"Çıxış {CHECKOUT_DEADLINE_HOUR}:00-dan sonra vurulmağa cəhd edildi. Cari vaxt: {now.strftime('%H:%M')}" ,
                )
            return

        if action == "checkin":
            # Qayda 1: GPS aktivdir? (Koordinatlar düzgündürmü?)
            # Location göndərilmişsə, GPS aktivdir, amma koordinatların düzgün olduğunu yoxlayırıq
            if lat == 0.0 and lon == 0.0:
                prof = db.get_user_by_telegram_id(user_id)
                name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
                user_phone = prof.get("phone_number") if prof else None
                
                await message.answer(
                    "❌ GPS koordinatları düzgün deyil. GPS-i aktiv edin və yenidən cəhd edin.",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="GPS problemi",
                        details=f"Giriş zamanı GPS koordinatları düzgün deyil (0.0, 0.0)"
                    )
                return
            
            # Qayda 2: Bu gün artıq giriş vurulub?
            existing = db.get_user_session_on_date(uid, today)
            if existing:
                await message.answer("❌ Bu gün artıq giriş etmisiniz. Giriş-çıxış yalnız bir dəfə vurula bilər.", reply_markup=worker_keyboard())
                pending_action.pop(user_id, None)
                return
            
            # Qayda 3: Giriş 11:00-a qədər vurulmalıdır
            if now.hour >= CHECKIN_DEADLINE_HOUR:
                prof = db.get_user_by_telegram_id(user_id)
                name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
                user_phone = prof.get("phone_number") if prof else None
                
                await message.answer(
                    f"❌ Giriş {CHECKIN_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
                    f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="Gecikmə - giriş vaxtı keçib",
                        details=f"Giriş {CHECKIN_DEADLINE_HOUR}:00-dan sonra vurulmağa cəhd edildi. Cari vaxt: {now.strftime('%H:%M')}"
                    )
                return
            
            # Qayda 4 (dəyişdirildi): Giriş zamanı məkan məhdudiyyəti tətbiq edilmir.

            db.create_session(user_id=uid, start_time=now_iso, lat=lat, lon=lon)

            start_link = f"https://maps.google.com/?q={lat},{lon}"
            kb = InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="Xəritədə bax", url=start_link)]]
            )
            prof = db.get_user_by_telegram_id(user_id)
            name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
            code = prof.get("code") if prof else None
            
            # Cavabı dərhal göndər (adres yüklənməsini gözləmə)
            info_lines = [
                "✅ Giriş qeyd olundu",
                f"👤 {name}" + (f" | Kod: {code}" if code else ""),
                f"📅 {today}  ⏰ {now.strftime('%H:%M:%S')}",
                f"📍 Koordinatlar: {lat}, {lon}",
            ]
            await message.answer("\n".join(info_lines))
            await message.answer("📍 Başlanğıc nöqtəsi", reply_markup=kb)
            
            # Xatırlatma mesajı
            reminder_text = (
                "💡 Xatırlatma\n\n"
                "Gününüz xoş keçsin! 🌟\n\n"
                "⚠️ Çıxış etməyi unutmayın!\n"
                "Yoxsa iş günü kimi hesablanmayacaq."
            )
            await message.answer(reminder_text)
            
            await message.answer("Menyu:", reply_markup=worker_keyboard())

            # Legacy attendance (save without address initially)
            try:
                if prof and isinstance(prof.get("id"), int):
                    legacy_user_id = int(prof["id"])  # type: ignore[index]
                    db.record_giris(
                        user_id=legacy_user_id,
                        date=today,
                        time=now.strftime("%H:%M:%S"),
                        location=None,  # Will be updated by background task if geocoding enabled
                    )
            except Exception as e:
                print(f"[handle_location checkin legacy] {e}")
            
            # Background geocoding task (non-blocking)
            async def send_address():
                try:
                    addr = await reverse_geocode(lat, lon)
                    if addr:
                        await message.answer(f"📍 Ünvan: {addr}")
                        # Update legacy attendance with address
                        if prof and isinstance(prof.get("id"), int):
                            legacy_user_id = int(prof["id"])
                            conn = db.sqlite3.connect(db.DB_FILE)
                            cursor = conn.cursor()
                            cursor.execute(
                                'UPDATE attendance SET giris_loc = ? WHERE user_id = ? AND date = ?',
                                (addr, legacy_user_id, today)
                            )
                            conn.commit()
                            conn.close()
                except Exception as e:
                    print(f"[send_address checkin] {e}")
            
            asyncio.create_task(send_address())

            # Schedule reminder after 8 hours
            asyncio.create_task(schedule_checkout_reminder(user_id, now, uid))

            pending_action.pop(user_id, None)
            return

        if action == "checkout":
            prof = db.get_user_by_telegram_id(user_id)
            name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
            user_phone = prof.get("phone_number") if prof else None
            
            # Qayda 1: GPS aktivdir? (Koordinatlar düzgündürmü?)
            if lat == 0.0 and lon == 0.0:
                await message.answer(
                    "❌ GPS koordinatları düzgün deyil. GPS-i aktiv edin və yenidən cəhd edin.",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="GPS problemi",
                        details=f"Çıxış zamanı GPS koordinatları düzgün deyil (0.0, 0.0)"
                    )
                return
            
            # Qayda 2: Giriş vurulubmu?
            sess = db.get_open_session(uid)
            if not sess:
                await message.answer(
                    "❌ Giriş etmədiyiniz üçün çıxış edə bilmirsiniz. Əvvəlcə giriş edin.",
                    reply_markup=worker_keyboard(),
                )
                pending_action.pop(user_id, None)
                return
            
            # Bu gün üçün artıq çıxış vurulubmu yoxlayırıq
            today_sess = db.get_user_session_on_date(uid, today)
            if today_sess and today_sess.get("end_time"):
                await message.answer(
                    "❌ Bu gün artıq çıxış etmisiniz. Giriş-çıxış yalnız bir dəfə vurula bilər.",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                return

            # Compute metrics
            start_time = parse_dt_to_baku(sess["start_time"])  # type: ignore[index]
            duration_hours = (now - start_time).total_seconds() / 3600.0
            duration_min = max(0, int((now - start_time).total_seconds() // 60))
            
            start_lat = float(sess["start_lat"])  # type: ignore[index]
            start_lon = float(sess["start_lon"])  # type: ignore[index]
            dist_m = haversine_m(start_lat, start_lon, lat, lon)
            
            # Qayda 3: Çıxış 19:00-a qədər vurulmalıdır
            if now.hour >= CHECKOUT_DEADLINE_HOUR:
                await message.answer(
                    f"❌ Çıxış {CHECKOUT_DEADLINE_HOUR}:00-dan sonra vurula bilməz. "
                    f"Hal-hazırda vaxt: {now.strftime('%H:%M')}",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="Gecikmə - çıxış vaxtı keçib",
                        details=f"Çıxış {CHECKOUT_DEADLINE_HOUR}:00-dan sonra vurulmağa cəhd edildi. Cari vaxt: {now.strftime('%H:%M')}"
                    )
                return
            
            # Qayda 4: Minimum 3 saat keçibmi?
            if duration_hours < MIN_WORK_DURATION_HOURS:
                await message.answer(
                    f"❌ Ən azı {MIN_WORK_DURATION_HOURS} saat sonra çıxış edə bilərsiniz.\n\n"
                    f"⏱ Hal-hazırda keçən vaxt: {duration_min} dəqiqə ({duration_hours:.1f} saat)\n"
                    f"📅 Giriş vaxtı: {start_time.strftime('%H:%M')}",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="Minimum iş müddəti pozulub",
                        details=f"Girişdən sonra yalnız {duration_hours:.1f} saat keçib. Minimum: {MIN_WORK_DURATION_HOURS} saat. Giriş vaxtı: {start_time.strftime('%H:%M')}"
                    )
                return
            
            # Qayda 5 (praktika): Giriş-çıxış eyni nöqtə məhdudiyyəti tətbiq edilmir.
            
            # Qayda 6: Məkəndədir?
            dist_from_workplace = haversine_m(WORKPLACE_LAT, WORKPLACE_LON, lat, lon)
            if dist_from_workplace > WORKPLACE_RADIUS_M:
                await message.answer(
                    f"❌ Çıxış məkandan kənar vurula bilməz.\n\n"
                    f"📍 Məkan radiusu: {WORKPLACE_RADIUS_M} metr\n"
                    f"📍 Sizin məsafəniz: {int(dist_from_workplace)} metr\n\n"
                    f"Zəhmət olmasa məkan daxilində olun.",
                    reply_markup=worker_keyboard()
                )
                pending_action.pop(user_id, None)
                
                # Çağrı mərkəzinə bildiriş
                if ADMIN_ID != 0:
                    await notifications.notify_rule_violation(
                        bot=bot,
                        admin_id=ADMIN_ID,
                        user_id=user_id,
                        user_name=name,
                        user_phone=user_phone,
                        violation_type="Məkandan kənar çıxış cəhdi",
                        details=f"Lokasiya: {lat}, {lon}. Məsafə: {int(dist_from_workplace)} metr (Maksimum: {WORKPLACE_RADIUS_M} metr)"
                    )
                return
            db.close_session(
                session_id=int(sess["id"]),  # type: ignore[index]
                end_time=now_iso,
                end_lat=lat,
                end_lon=lon,
                duration_min=duration_min,
                distance_m=dist_m,
            )

            start_link = f"https://maps.google.com/?q={start_lat},{start_lon}"
            end_link = f"https://maps.google.com/?q={lat},{lon}"
            route_link = (
                "https://www.google.com/maps/dir/?api=1"
                f"&origin={start_lat},{start_lon}"
                f"&destination={lat},{lon}"
                "&travelmode=walking"
            )
            ikb = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(text="Başlanğıc", url=start_link),
                        InlineKeyboardButton(text="Bitiş", url=end_link),
                    ],
                    [InlineKeyboardButton(text="Marşrut", url=route_link)],
                ]
            )

            prof = db.get_user_by_telegram_id(user_id)
            name = (prof.get("name") if prof else user.full_name) or "Istifadəçi"
            code = prof.get("code") if prof else None
            
            # Cavabı dərhal göndər
            info_lines = [
                "✅ Çıxış qeyd olundu",
                f"👤 {name}" + (f" | Kod: {code}" if code else ""),
                f"📅 {today}  ⏰ {now.strftime('%H:%M:%S')}",
                f"📍 Koordinatlar: {lat}, {lon}",
                f"⏱ İş müddəti: {duration_min} dəqiqə ({duration_hours:.1f} saat)",
            ]
            await message.answer("\n".join(info_lines))
            await message.answer("🗺 Xəritə linkləri", reply_markup=ikb)
            
            # Uğurlu tamamlanma mesajı
            success_message = (
                "🎉 İş gününüz uğurla tamamlandı!\n\n"
                "Xoş istirahətlər! 😊"
            )
            await message.answer(success_message)
            
            await message.answer("Menyu:", reply_markup=worker_keyboard())

            # Legacy attendance
            try:
                if prof and isinstance(prof.get("id"), int):
                    legacy_user_id = int(prof["id"])  # type: ignore[index]
                    db.record_cixis(
                        user_id=legacy_user_id,
                        date=today,
                        time=now.strftime("%H:%M:%S"),
                        location=None,
                    )
            except Exception as e:
                print(f"[handle_location checkout legacy] {e}")
            
            # Background geocoding task (non-blocking)
            async def send_address():
                try:
                    end_addr = await reverse_geocode(lat, lon)
                    if end_addr:
                        await message.answer(f"📍 Ünvan: {end_addr}")
                        # Update legacy attendance with address
                        if prof and isinstance(prof.get("id"), int):
                            legacy_user_id = int(prof["id"])
                            conn = db.sqlite3.connect(db.DB_FILE)
                            cursor = conn.cursor()
                            cursor.execute(
                                'UPDATE attendance SET cixis_loc = ? WHERE user_id = ? AND date = ?',
                                (end_addr, legacy_user_id, today)
                            )
                            conn.commit()
                            conn.close()
                except Exception as e:
                    print(f"[send_address checkout] {e}")
            
            asyncio.create_task(send_address())

            pending_action.pop(user_id, None)
            return

    except Exception as e:
        print(f"[handle_location] error: {e}")
        await message.answer("❌ Xəta baş verdi. Yenidən yoxlayın.", reply_markup=worker_keyboard())
        if message.from_user:
            pending_action.pop(message.from_user.id, None)
        return


# ================== FALLBACK HANDLER ==================

@dp.message()
async def fallback_show_menu(message: Message, state: FSMContext) -> None:
    cur = await state.get_state()
    if cur:
        return
    user = message.from_user
    if not user:
        return
    if is_admin(user.id):
        await message.answer("Admin menyusundan istifadə edin:", reply_markup=admin_keyboard())
    else:
        await message.answer("Menyudan istifadə edin:", reply_markup=worker_keyboard())


# ================== MAIN ==================

async def main() -> None:
    print("aiogram bot starting (GPS attendance)...")

    if not acquire_single_instance_lock():
        print(f"Bot artıq işləyir (lokal lock): {LOCK_FILE_PATH}. Əgər əminsiniz ki bot işləmir, bu faylı silin.")
        return
    atexit.register(release_single_instance_lock)

    try:
        await bot.delete_webhook(drop_pending_updates=True)
    except Exception:
        pass

    commands: list[BotCommand] = []
    await bot.set_my_commands(commands)

    # Admin üçün xüsusi əmrlər yoxdur - yalnız /start istifadə edir və butonlardan idarə edir
    if ADMIN_ID != 0:
        try:
            # Admin üçün yalnız start əmri qoyuruq
            admin_commands: list[BotCommand] = []
            await bot.set_my_commands(admin_commands, scope=BotCommandScopeChat(ADMIN_ID))
        except Exception:
            # If setting scoped commands fails, ignore so bot still works
            pass

    # Initialize PostgreSQL connection pool (must be before any DB operations)
    db.initialize_pool()

    # Ensure DB schema exists before handling any updates
    db.init_db()
    db.init_gps_tables()
    db.init_group_codes()
    db.init_registrations()

    try:
        await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
    except TelegramConflictError:
        print("TelegramConflictError: eyni BOT_TOKEN ilə başqa bot instansiyası işləyir. Digər prosesi/dayployment-i dayandırın.")
    finally:
        # Clean up connection pool on shutdown
        db.close_pool()
        print("✓ Database connection pool closed")


if __name__ == "__main__":
    asyncio.run(main())
